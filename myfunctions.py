import pandas as pd
import numpy as np
from append_df_to_excel import append_df_to_excel
from append_df_to_excel_existing import append_df_to_excel_existing
from datetime import datetime
import datetime as dt
from openpyxl import Workbook
import openpyxl
import re
import os
import PySimpleGUI as sg
import statistics
import math
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from scipy.optimize import curve_fit
import calendar
import sys
import string

# Dataframe creation /File reading -------------------------------------------------------------------------------------

def activetrackerdf(df_15m, df_tracker_info_calc):
    '''This script  creates a dataframe containing all active tracker incidents from a given report.
    It doesn't separate the data by sites. All data is gathered in the same dataframe

    Returns df_tracker_active'''

    for index, row in df_15m.iterrows():
        rel_comp = df_15m.loc[index, 'Related Component']
        state = df_15m.loc[index, 'State']
        site = df_15m.loc[index, 'Site Name']
        site = correct_site_name(site)

        if 'Tracker' in rel_comp or "TRACKER" in rel_comp:
            # print(rel_comp)
            if not 'gateway' in rel_comp:
                df_15m.loc[index, 'Tracker'] = 'Yes'
                df_15m.loc[index, 'Capacity Related Component'] = df_tracker_info_calc.loc[
                    site, 'Avg. capacity per tracker (kW)']
                """if 'Tracker B' in rel_comp:
                    tracker_group = re.search(r'B\d\d\d', rel_comp)
                    tracker_group = tracker_group.group()
                    df_15m.loc[index, 'Comments'] = df_tracker_impact.loc[tracker_group, 'Block']"""
        elif 'Tracker target availability' in state:
            df_15m.loc[index, 'Tracker'] = 'Yes'
            df_15m.loc[index, 'Capacity Related Component'] = df_tracker_info_calc.loc[
                site, 'Avg. capacity per tracker (kW)']

    df_15m_tracker = df_15m.loc[df_15m['Tracker'] == 'Yes']
    df_tracker_active = df_15m_tracker.loc[pd.isna(df_15m_tracker['Event End Time'])]
    df_tracker_active = df_tracker_active.reset_index(None, drop=True)

    '''for index, row in df_tracker_active.iterrows():
        time = df_tracker_active.loc[index, 'Event Start Time']
        if type(time) is datetime:
            df_tracker_active.loc[index, 'Event Start Time'] = str(time)
        if "." in time:
            df_tracker_active.loc[index, 'Event Start Time'] = str(time[:-4])'''

    return df_tracker_active

def create_df_list(df):
    '''This script creates two dictionaries and a list.
     The dictionaries contain all sites from a given report and their active and closed events dataframes
     The list contains all sites present in the report

     Returns site_list, df_list_active, df_list_closed'''

    site_list = df['Site Name']
    site_list = site_list.drop_duplicates()
    site_list = site_list.tolist()

    for site in site_list:
        index_site = site_list.index(site)
        site = correct_site_name(site)
        site_list[index_site] = site

    for site in site_list:
        if "LSBP - " in site or "LSBP – " in site:
            onlysite = site[7:]
        else:
            onlysite = site
        onlysite = onlysite.replace(" ", "")
        df_name_active = "df_15m_" + onlysite + "_active"
        df_name_closed = "df_15m_" + onlysite + "_closed"

        try:
            if site not in df_list_active.keys():
                df_list_active[site] = df_name_active

        except NameError:
            df_list_active = {site: df_name_active}

        try:
            if site not in df_list_closed.keys():
                df_list_closed[site] = df_name_closed

        except NameError:
            df_list_closed = {site: df_name_closed}
    print(site_list)
    return site_list, df_list_active,df_list_closed

def create_active_events_df(df):

    '''This script creates a dataframe containing all active event from a given report
    It doesn't filter the incidents by site, gathering them all in a single dataframe

    Returns df_active'''

    df_15m_notprod = df.loc[df['Component Status'] == 'Not Producing']
    df_active = df_15m_notprod.loc[df_15m_notprod['Event End Time'].isnull()]

    return df_active

def closedtrackerdf(df_15m, df_tracker_info_calc):
    '''This script  creates a dataframe containing all active tracker incidents from a given report.
    It doesn't separate the data by sites. All data is gathered in the same dataframe
    This script doesn't filter the incidents by duration.

    Returns df_15m_tracker_active'''

    for index, row in df_15m.iterrows():
        rel_comp = df_15m.loc[index, 'Related Component']
        state = df_15m.loc[index, 'State']
        site = df_15m.loc[index, 'Site Name']
        site = correct_site_name(site)

        if 'Tracker' in rel_comp or "TRACKER" in rel_comp:
            # print(rel_comp)
            if not 'gateway' in rel_comp:
                df_15m.loc[index, 'Tracker'] = 'Yes'
                df_15m.loc[index, 'Capacity Related Component'] = df_tracker_info_calc.loc[
                    site, 'Avg. capacity per tracker (kW)']
                """if 'Tracker B' in rel_comp:
                    tracker_group = re.search(r'B\d\d\d', rel_comp)
                    tracker_group = tracker_group.group()
                    df_15m.loc[index, 'Comments'] = df_tracker_impact.loc[tracker_group, 'Block']"""
        elif 'Tracker target availability' in state:
            df_15m.loc[index, 'Tracker'] = 'Yes'
            df_15m.loc[index, 'Capacity Related Component'] = df_tracker_info_calc.loc[
                site, 'Avg. capacity per tracker (kW)']

    df_15m_tracker = df_15m.loc[df_15m['Tracker'] == 'Yes']
    df_15m_tracker_closed = df_15m_tracker.loc[df_15m_tracker['Event End Time'].notnull()]
    df_15m_tracker_dur15m = df_15m_tracker_closed.loc[df_15m_tracker_closed['Duration (h)'] > 0.249]  # filter by duration >15 minute
    df_15m_tracker_final = df_15m_tracker_dur15m.reset_index(None, drop=True)  # reset index

    '''for index, row in df_15m_tracker_final.iterrows():
        time = df_15m_tracker_final.loc[index, 'Event Start Time']
        if type(time) is datetime:
            df_15m_tracker_final.loc[index, 'Event Start Time'] = str(time)
            time = df_15m_tracker_final.loc[index, 'Event Start Time']
        if ".0" in time:
            df_15m_tracker_final.loc[index, 'Event Start Time'] = str(time[:-4])

    for index, row in df_15m_tracker_final.iterrows():
        time = df_15m_tracker_final.loc[index, 'Event End Time']
        if type(time) is datetime:
            df_15m_tracker_final.loc[index, 'Event End Time'] = str(time)
        if "." in time:
            df_15m_tracker_final.loc[index, 'Event End Time'] = str(time[:-4])'''

    return df_15m_tracker_final

def read_Daily_Alarm_Report(Alarm_report_path, irradiance_file_path, event_tracker_path, previous_dmr_path):
    dir = os.path.dirname(Alarm_report_path)
    basename = os.path.basename(Alarm_report_path)
    date_finder = re.search(r'\d\d\d\d-\d\d-\d\d', Alarm_report_path)
    date = date_finder.group()
    geography_report_match = re.search(r'\w+?_', basename)
    geography_report = geography_report_match.group()[:-1]
    print(geography_report)

    day = date[-2:]     #so naive please change this later
    month = date[-5:-3]
    year = date[:4]


    df_all = pd.read_excel(Alarm_report_path, engine="openpyxl")
    df_all['InSolar Check'] = ""
    df_all['Curtailment Event'] = ""
    df_all['Tracker'] = ''
    df_all['Comments'] = ''
    df_all_columns = df_all.columns

    irradiance_data = pd.read_excel(irradiance_file_path, engine="openpyxl")

    try:
        all_prev_active_events = pd.read_excel(previous_dmr_path,sheet_name=["Active Events","Active tracker incidents"], engine="openpyxl")
        #all_prev_active_events = pd.concat([all_prev_active_events['Active Events'], all_prev_active_events['Active tracker incidents']])
        #df_all = pd.concat([df_all, all_prev_active_events['Active Events'], all_prev_active_events['Active tracker incidents']])[df_all_columns]

        prev_active_events = all_prev_active_events['Active Events']
        prev_active_tracker_events = all_prev_active_events['Active tracker incidents']

        print(df_all.columns)
    except FileNotFoundError:
        print("Previous Daily Monitoring Report not found.")
        try:
            all_prev_active_events = pd.read_excel(event_tracker_path,sheet_name=["Active Events","Active tracker incidents"], engine="openpyxl")
            #all_prev_active_events = pd.concat([all_prev_active_events['Active Events'], all_prev_active_events['Active tracker incidents']])

            prev_active_events = all_prev_active_events['Active Events']
            prev_active_tracker_events = all_prev_active_events['Active tracker incidents']

        except FileNotFoundError:
            print("Event Trakcer not found.")


    newfile = dir + '/Incidents' + str(day) + '-' + str(month) + str(geography_report) + '.xlsx'
    newtrackerfile = dir + '/Tracker_Incidents' + str(day) + '-' + str(month) + str(geography_report) + '.xlsx'


    return df_all, newfile, newtrackerfile, irradiance_data, prev_active_events, prev_active_tracker_events

def read_general_info(Report_template_path, general_info_path):

    df_general_info = pd.read_excel(general_info_path, sheet_name = 'Site Info', engine="openpyxl")
    df_general_info_calc = pd.read_excel(general_info_path, sheet_name = 'Site Info', index_col=0, engine="openpyxl")

    #df_tracker_impact = pd.read_excel(trackers_info_path, sheet_name='Info_Trackers_Impact', engine="openpyxl",
                                      #index_col=0)
    all_component_data = pd.read_excel(general_info_path, sheet_name='Component Code', index_col=0, engine="openpyxl")

    return df_general_info, df_general_info_calc,all_component_data

def fill_events_analysis_dataframe(df_analysis, df_info_sunlight):
    max_percentage = "{:.2%}".format(1)
    site_list = df_info_sunlight['Site']
    for site in site_list:
        index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site].index.values
        index_site = int(index_site_array[0])
        stime = df_info_sunlight.loc[index_site, 'Time of operation start']
        etime = df_info_sunlight.loc[index_site, 'Time of operation end']

        index_mint = df_analysis[
            df_analysis['Time'] == stime].index.values  # gets starting time row index
        int_index_mint = index_mint[0]  # turns index from numpy.ndarray to integer

        index_maxt = df_analysis[
            df_analysis['Time'] == etime].index.values  # gets ending time row index
        int_index_maxt = index_maxt[0]

        for index in range(int_index_mint, int_index_maxt):
            df_analysis.loc[index, site] = max_percentage

    return df_analysis

def get_filename_folder():
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.

    layout = [[sg.Text('Choose file', pad=((2, 10), (2, 5)))],
              [sg.FileBrowse(target='-FILE-'),
               sg.In(key='-FILE-', text_color='black', size=(20, 1), enable_events=True, readonly=True,
                     visible=True)],
              [sg.Button('Submit'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Choose file', layout)

    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read(timeout=100)

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            "file_path = filename = folder = 0"
            break

        if event == 'Submit':
            file_path = values['-FILE-']
            print(file_path)
            if file_path == "":
                sg.popup('No file chosen, try again or exit')
                continue
            filepath_split = os.path.split(file_path)
            folder = filepath_split[0] + "/"
            filename = filepath_split[1].split(".")[0]
            extension = "." + filepath_split[1].split(".")[1]

            break




    window.close()


    return file_path,filename,folder,extension







# Structure Data -------------------------------------------------------------------------------------------------------
def match_df_to_event_tracker(df, component_data,fmeca_data, active: bool = False , simple_match: bool = False):
    desired_columns_components = ["ID", "Site Name", "Related Component", "Capacity Related Component",
                                  "Component Status", "Event Start Time","Event O&M Response Time", "Event End Time",
                                  "Duration (h)","Active Hours (h)", "Energy Lost (MWh)","Comments", "Remediation",
                                  "Fault","Fault Component","Failure Mode", "Failure Mechanism",
                                  "Category","Subcategory","Resolution Category","Incident Status", "Categorization Status"]
    if simple_match == False:
        curtailment_fmeca = fmeca_data.loc[fmeca_data['Failure Mode'] == 'Curtailment']

        df['Site Name'] = [correct_site_name(name) for name in df['Site Name']]
        df['Related Component'] = [correct_site_name(name) for name in df['Related Component']]

        #print(df['Site Name'],df['Related Component'])

        #Add ID column
        if not "ID" in df.columns:
            df = change_times_to_str(df, active=active)

            # test if all entries to add have an entry on the general info file
            tuple_list_toadd = list(set([(row['Site Name'], row['Related Component']) for index, row in df.iterrows()]))
            tuple_list_componentdata = [(row['Site'], row['Component']) for index, row in component_data.iterrows()]

            does_not_exist = [x for x in tuple_list_toadd if x not in tuple_list_componentdata]

            if not does_not_exist:
                df.insert(0, "ID", [component_data.loc[(component_data['Site'] == df.loc[index, 'Site Name']) & (component_data['Component'] == df.loc[index, 'Related Component'])]['ID'].values[0] + '-' + df.loc[index,'Event Start Time'].replace(" ", "T").replace("-", "").replace(":", "")for index, row in df.iterrows()])
            else:
                print(does_not_exist)
                print(df.loc[df['Related Component'] == does_not_exist[1]])
                """print(tuple_list_toadd)
                print(tuple_list_componentdata)"""
                sys.exit("These components do not exist in the general info file")
            """for index, row in df.iterrows():
                test = component_data.loc[(component_data['Site'] == df.loc[index, 'Site Name']) & (component_data['Component'] == df.loc[index, 'Related Component'])]['ID'].values[0]
                time_test = df.loc[index,'Event Start Time'].replace(" ", "T").replace("-", "").replace(":", "")
                print(row['Site Name'], row['Related Component'], test, time_test)"""



    # Add rest of the columns
    for column in desired_columns_components:
        if not column in df.columns:
            df[column] = ""
        elif column == 'Incident Status':
            if active == True:
                df[column] = "Open"
            else:
                df[column] = "Closed"
        elif column == 'Categorization Status':
            if active == True:
                df[column] = "Pending"
            else:
                df[column] = ["Pending" if status == "" else status for status in df[column]]


    df.drop_duplicates(subset = ['ID'], inplace = True, ignore_index = True)#.reset_index(drop=True, inplace=True)

    if simple_match == False:
        for index, row in df.loc[df['Curtailment Event'] == 'x'].iterrows():
            df.loc[index, 'Fault'] = curtailment_fmeca['Fault'].values[0]
            df.loc[index, 'Fault Component'] = curtailment_fmeca['Fault Component'].values[0]
            df.loc[index, 'Failure Mode'] = curtailment_fmeca['Failure Mode'].values[0]
            df.loc[index, 'Failure Mechanism'] = curtailment_fmeca['Failure Mechanism'].values[0]
            df.loc[index, 'Category'] = curtailment_fmeca['Category'].values[0]
            df.loc[index, 'Subcategory'] = curtailment_fmeca['Subcategory'].values[0]
            df.loc[index, 'Resolution Category'] = "Reset"
            df.loc[index, 'Incident Status'] = "Closed"
            df.loc[index, 'Categorization Status'] = "Completed"

    df.drop(columns=df.columns.difference(desired_columns_components), inplace=True)

    df_final = df[desired_columns_components]

    return df_final



# Input Data -----------------------------------------------------------------------------------------------------------

def input_date(startend: str = "start"):

    hour = [*range(24)]
    for i in range(0, len(hour), 1):
        hour[i] = str(f'{hour[i]:02}')

    minutes = [*range(0, 46, 15)]
    for i in range(0, len(minutes), 1):
        minutes[i] = str(f'{minutes[i]:02}')

    # Create interface
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.
    layout = [[sg.Text('Enter date of report you want to analyse')],
              [sg.CalendarButton('Choose ' + startend + ' date', target='-CAL-', format="%Y-%m-%d"),
               sg.In(key='-CAL-', text_color='black', size=(10, 1), enable_events=True, readonly=True, visible=True)],
              [sg.Button('Submit'), sg.Exit()]]
    # Create the Window
    window = sg.Window('Choose date', layout)
    # Event Loop to process "events" and get the "values" of the inputs

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            break
        if event == 'Submit':
            date = values['-CAL-']
            break
    window.close()


    return date

def input_date_and_time():
    hour = [*range(24)]
    for i in range(0, len(hour), 1):
        hour[i] = str(f'{hour[i]:02}')

    minutes = [*range(0, 46, 15)]
    for i in range(0, len(minutes), 1):
        minutes[i] = str(f'{minutes[i]:02}')

    # Create interface
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.
    layout = [[sg.Text('Enter date of report you want to analyse')],
              [sg.CalendarButton('Choose date that the event started', target='-CAL-', format="%Y-%m-%d"),
               sg.In(key='-CAL-', text_color='black', size=(10, 1), enable_events=True, readonly=True, visible=True)],
              [sg.Text('Enter start time of event'), sg.Spin(hour, initial_value='07',size = (3,2), key='-SHOUR-'),
               sg.Spin(minutes, initial_value='00',size = (3,2), key='-SMIN-')],
              [sg.Button('Submit'), sg.Exit()]]
    # Create the Window
    window = sg.Window('Daily Monitoring Report', layout)
    # Event Loop to process "events" and get the "values" of the inputs

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            break
        if event == 'Submit':
            date = values['-CAL-']

            stime_hour = values['-SHOUR-']
            stime_min = values['-SMIN-']

            stime = date + ' ' + stime_hour + ':' + stime_min + ':00'
            timestamp = datetime.strptime(stime, '%Y-%m-%d %H:%M:%S')

            print(timestamp)
            break
    window.close()


    return timestamp

def input_time_operation_site(site, date):

    hour = [*range(24)]
    for i in range(0, len(hour),1):
        hour[i] = str(f'{hour[i]:02}')

    minutes = [*range(0,46,15)]
    for i in range(0, len(minutes),1):
        minutes[i] = str(f'{minutes[i]:02}')


    #Create interface
    sg.theme('DarkAmber')  # Add a touch of color
        # All the stuff inside your window.
    layout = [[sg.Text('Enter sunrise and sunset time for ' + site)],
              [sg.HorizontalSeparator(pad=((10, 10), (2, 10)))],
              [sg.Text('Enter sunrise hour'), sg.Spin(hour, initial_value = '07',size = (3,2), key = '-SHOUR-'),
           sg.Spin(minutes, initial_value = '00',size = (3,2), key = '-SMIN-')],
          [sg.Text('Enter sunset hour'), sg.Spin(hour, initial_value= '19',size = (3,2), key='-EHOUR-'),
           sg.Spin(minutes, initial_value= '00',size = (3,2), key='-EMIN-')],
              [sg.Button('Submit'), sg.Exit()]]
    # Create the Window
    window = sg.Window('Daily Monitoring Report', layout)
    # Event Loop to process "events" and get the "values" of the inputs

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            break
        if event == 'Submit':
            stime_hour = values['-SHOUR-']
            stime_min = values['-SMIN-']

            print(date, type(date))

            stime = date + ' ' + stime_hour + ':' + stime_min + ':00'
            stime = datetime.strptime(stime, '%Y-%m-%d %H:%M:%S')

            etime_hour = values['-EHOUR-']
            etime_min = values['-EMIN-']

            etime = date + ' ' + etime_hour + ':' + etime_min + ':00'
            etime = datetime.strptime(etime, '%Y-%m-%d %H:%M:%S')

            print(stime)
            print(etime)
            break
    window.close()

    return stime, etime




# Data Treatment and Cleansing----------------------------------------------------------------------------------------------

def correct_incidents_irradiance_for_overlapping_parents(incidents, irradiance, component_data,recalculate: bool = False, timestamp: float = 15,
                                                         irradiance_threshold: float = 20):
    '''From: Incidents table, irradiance
    Returns: Dict with Irradiance dataframe corrected for overlapping parents events, i.e.,
    removes periods where parents incidetns are active. For each incident'''

    incidents_corrected_info = {}
    granularity = timestamp / 60
    granularity_str = str(timestamp) + "min"

    if recalculate == True:
        pass
    else:
        n_inc_1 = incidents.shape[0]
        incidents = incidents.loc[(incidents['Active Hours (h)'].isna()) | (incidents['Event End Time'].isna())]
        n_inc_2 = incidents.shape[0]
        print('No recalculation, analysing ', n_inc_2, ' from a total of ', n_inc_1, ' incidents.' )


    for index, row in incidents.iterrows():

        # Get site info
        site = row['Site Name']
        site_info = component_data.loc[component_data['Site'] == site]
        site_capacity = float(component_data.loc[component_data['Component'] == site]['Nominal Power DC'].values)
        'budget_pr_site = budget_pr.loc[site, :]'

        # Get site Incidents
        site_incidents = incidents.loc[incidents['Site Name'] == site]

        # Get site irradiance
        df_irradiance_site = irradiance.loc[:, irradiance.columns.str.contains(site + '|Timestamp')]

        # Get irradiance poa avg column
        poa_avg_column = df_irradiance_site.loc[:, df_irradiance_site.columns.str.contains('Irradiance')]
        poa_avg_column = poa_avg_column.loc[:, ~poa_avg_column.columns.str.contains('curated')].columns.values[0]

        """# Get first timestamp under analysis and df from that timestamp onwards
        stime_index = next(i for i, v in enumerate(df_irradiance_site[poa_avg_column]) if v > irradiance_threshold)
        site_start_time = df_irradiance_site['Timestamp'][stime_index]

        df_irradiance_operation_site = df_irradiance_site.loc[df_irradiance_site['Timestamp'] >= site_start_time]
        df_irradiance_operation_site['Day'] = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S').date() for
                                               timestamp in
                                               df_irradiance_operation_site['Timestamp']]"""

        # Incident Info
        id_incident = row['ID']
        capacity = row['Capacity Related Component']
        parents = (site_info.loc[site_info['Component'] == row['Related Component']]).loc[:,
                  site_info.columns.str.contains('Parent')].values.flatten().tolist()
        parents = [x for x in parents if str(x) != 'nan']

        try:
            # Active Events-------------------------------------------------------------------------------------
            math.isnan(row['Event End Time'])
            effective_start_time_incident = row['Event Start Time']

            # In active events, end time of incident is the latest record of irradiance
            effective_end_time_incident = datetime.strptime(str(df_irradiance_site['Timestamp'].to_list()[-1]),
                                                            '%Y-%m-%d %H:%M:%S')
            closed_event = False

        except TypeError:
            # Closed Events
            closed_event = True
            effective_start_time_incident = row['Event Start Time']
            effective_end_time_incident = row['Event End Time']

        finally:
            if len(parents) == 0:
                pass
            else:
                parents_incidents = site_incidents[site_incidents['Related Component'].isin(parents)]
                if not parents_incidents.empty:
                    relevant_parents_incidents = parents_incidents.loc[
                        ~(parents_incidents['Event End Time'] <= effective_start_time_incident) & ~(
                            parents_incidents['Event Start Time'] >= effective_end_time_incident)]
                    # relevant_parents_incidents = mf.rounddatesclosed_15m(site,relevant_parents_incidents)
                    if not relevant_parents_incidents.empty:
                        # cycle through parents incidents and get active hours to remove
                        # Get first timestamp under analysis and df from that timestamp onwards
                        irradiance_incident = df_irradiance_site.loc[
                            (df_irradiance_site['Timestamp'] >= effective_start_time_incident) & (
                                    df_irradiance_site['Timestamp'] <= effective_end_time_incident)]
                        irradiance_incident['Day'] = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S').date() for
                                                      timestamp in irradiance_incident['Timestamp']]

                        try:
                            del timestamps_to_remove
                        except NameError:
                            pass
                        if closed_event == False:
                            print('Relevant overlapping Incidents for an active event')
                        else:
                            print('Relevant overlapping Incidents for a closed event')

                        print(row[['Related Component', 'Event Start Time', 'Event End Time']])
                        print(relevant_parents_incidents[['Related Component', 'Event Start Time', 'Event End Time']])

                        for index_rpi, row_rpi in relevant_parents_incidents.iterrows():

                            rpi_start_time = \
                                pd.Series(row_rpi['Event Start Time']).dt.round(granularity_str, 'shift_backward')[0]
                            rpi_actual_start_time = row_rpi['Event Start Time']

                            try:
                                rpi_end_time = \
                                    pd.Series(row_rpi['Event End Time']).dt.round(granularity_str, 'shift_forward')[0]
                                rpi_actual_end_time = row_rpi['Event End Time']
                            except AttributeError:
                                rpi_end_time = effective_end_time_incident
                                rpi_actual_end_time = effective_end_time_incident

                            timestamp_range = list(pd.date_range(start=rpi_start_time, end=rpi_end_time, freq='15min'))
                            actual_timestamp_range = list(pd.date_range(start=rpi_start_time, end=rpi_end_time, freq='1min'))

                            try:
                                timestamps_to_remove += timestamp_range
                                actual_timestamps_to_remove += actual_timestamp_range
                            except (NameError, AttributeError):
                                timestamps_to_remove = timestamp_range
                                actual_timestamps_to_remove = actual_timestamp_range

                            # print("Look Here: \n", timestamps_to_remove)

                        timestamps_to_remove = sorted(set(timestamps_to_remove))
                        actual_timestamps_to_remove = sorted(set(actual_timestamps_to_remove))
                        overlapped_time_1m = len(actual_timestamps_to_remove)/60

                        timestamps_to_keep = [timestamp for timestamp in irradiance_incident['Timestamp'].to_list() if
                                              timestamp not in timestamps_to_remove]



                        corrected_irradiance_incident = irradiance_incident.loc[
                            irradiance_incident['Timestamp'].isin(timestamps_to_keep)]

                        actual_column = corrected_irradiance_incident.loc[:,
                                        corrected_irradiance_incident.columns.str.contains('Average')]
                        actual_column = actual_column.loc[:, ~actual_column.columns.str.contains('curated')].columns.values[0]

                        cleaned_irradiance = corrected_irradiance_incident.dropna(subset=[actual_column])






                        try:
                            data_gaps_proportion = 1 - (len(cleaned_irradiance[actual_column]) / len(corrected_irradiance_incident[actual_column]))
                        except ZeroDivisionError:
                            print('Divison by zero: ',len(cleaned_irradiance[actual_column]), " /", len(corrected_irradiance_incident[actual_column]) )
                            data_gaps_proportion = 1

                        incidents_corrected_info[id_incident] = {
                            'Corrected Irradiance Incident': corrected_irradiance_incident,
                            'Cleaned Corrected Irradiance Incident': cleaned_irradiance,
                            "Time overlapped 1m": overlapped_time_1m,
                            'Data Gaps Proportion':data_gaps_proportion,
                            'Irradiance Column': actual_column,
                            'Irradiance Raw': irradiance_incident}


                        # print(irradiance_incident)
                    else:
                        print('No overlapping parents')
                        continue

                else:
                    print('No overlapping parents')
                    continue

    return incidents_corrected_info

def correct_duration_of_event(df):
    for index, row in df.iterrows():

        stime = df.loc[index, 'Event Start Time']
        stime = datetime.strptime(stime, '%Y-%m-%d %H:%M:%S')

        etime = df.loc[index, 'Event End Time']
        etime = datetime.strptime(etime, '%Y-%m-%d %H:%M:%S')

        duration = etime - stime
        duration = duration.total_seconds()/ 3600

        df.loc[index,'Duration (h)'] = round(duration,2)

    return df

def correct_site_name(site):

    while site[-1] == " ":
        site = site[:-1]
    while site[0] == " ":
        site = site[1:]

    return site

def filter_notprod_and_duration(df, duration):
    ''' Filters df by not producing events and given duration in minutes

     Returns df'''
    d_hours = duration / 60.1  # turns minutes into hours

    df_closed = df.loc[df['Event End Time'].notnull()]

    df_15m_notprod = df_closed.loc[df['Component Status'] == 'Not Producing']  # new dataframe with only the not producing incidents

    df_15m_notprod_dur15m = df_15m_notprod.loc[df_15m_notprod['Duration (h)'] > d_hours]  # filter by duration >X minutes

    df_final = df_15m_notprod_dur15m.reset_index(None, drop=True)  # resets index

    return df_final

def filter_bystartdate_component(df):

    # Create interface
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.
    layout = [[sg.Text('Choose dates betweem which to filter start time of event')],
              [sg.CalendarButton('Choose date', target='-CALS-',default_date_m_d_y = None, format="%Y-%m-%d"),
               sg.In(key='-CALS-', text_color='black', size=(10, 1), enable_events=True, readonly=True, visible=True)],
              [sg.CalendarButton('Choose date', target='-CALE-',default_date_m_d_y = None, format="%Y-%m-%d"),
               sg.In(key='-CALE-', text_color='black', size=(10, 1), enable_events=True, readonly=True, visible=True)],
              [sg.Button('Submit'), sg.Exit()]]
    # Create the Window
    window = sg.Window('Daily Monitoring Report', layout)
    # Event Loop to process "events" and get the "values" of the inputs

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            break
        if event == 'Submit':
            stime_initial = values['-CALS-']
            stime_final = values['-CALE-']

            print(stime_initial)
            print(stime_final)
            break
    window.close()


    stime_column = df['Event Start Time']

    stime_column = [datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S') for timestamp in stime_column]
    etime_column = [datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S') for timestamp in etime_column]

    df['Event Start Time'] = stime_column
    df['Event End Time'] = etime_column

    #df[(df['Event Start Time'] > '2014-07-23 07:30:00') & (df['Event Start Time'] < '2014-07-23 09:00:00')]

    return df

def get_actual_irradiance_column(df_irradiance_site):
    ''' From irradiance data of a site, containing the different types (curated, poa avg) selects curated if data gaps
    are less than 25%, if not it selects poa avergae if data gaps for that is less than 60%, if data gaps persist,
    returns no irradiance given that the results are invalid'''


    # Search for irradiance columns to use, if there is no curated irradiance, use POA average
    try:
        curated_column = df_irradiance_site.loc[:, df_irradiance_site.columns.str.contains('curated')].columns.values[
            0]
    except IndexError:
        curated_column = None
        print("no curated column")
    try:
        poa_avg_column = df_irradiance_site.loc[:, df_irradiance_site.columns.str.contains('Average')]
        poa_avg_column = poa_avg_column.loc[:, ~poa_avg_column.columns.str.contains('curated')].columns.values[0]
    except IndexError:
        poa_avg_column = None

    # print(curated_column)
    # print(poa_avg_column)

    # Data column check to determine what irradiance to use
    if curated_column:
        # Verify data gap percentage of curated column & then POA average
        cleaned_irradiance = [value for value in df_irradiance_site[curated_column] if not math.isnan(value)]
        #print(len(cleaned_irradiance),len(df_irradiance_site[curated_column]))
        try:
            data_gaps_proportion = 1 - (len(cleaned_irradiance) / len(df_irradiance_site[curated_column]))
        except ZeroDivisionError:
            data_gaps_proportion = 1

        if data_gaps_proportion > 0.1:

            # verify POA average
            try:
                cleaned_irradiance = [value for value in df_irradiance_site[poa_avg_column] if not math.isnan(value)]
                data_gaps_proportion = 1 - (len(cleaned_irradiance) / len(df_irradiance_site[poa_avg_column]))
            except ZeroDivisionError:
                data_gaps_proportion = 1

            if data_gaps_proportion > 0.60:
                actual_column = None

            else:
                actual_column = poa_avg_column


        else:
            actual_column = curated_column



    elif poa_avg_column:
        actual_column = poa_avg_column

        # Verify data gap percentage of poa avg column
        try:
            cleaned_irradiance = [value for value in df_irradiance_site[actual_column] if not math.isnan(value)]
            data_gaps_proportion = 1 - (len(cleaned_irradiance) / len(df_irradiance_site[actual_column]))

            if data_gaps_proportion > 0.60:
                actual_column = None
        except ZeroDivisionError:
            actual_column = None
            data_gaps_proportion = 1

    else:
        actual_column = None
        data_gaps_proportion = 1

    return actual_column,curated_column,data_gaps_proportion, poa_avg_column

def get_site_and_inverter_data_ready_for_analysis(inverters_data, component_data, budget_pr, general_info):

    inverter_list_raw = inverters_data.columns[inverters_data.columns.str.contains('AC')].to_list()
    inverter_list = [re.search(r'Inverter \d.+', name).group().replace(']', "") for name in inverter_list_raw]

    irradiance_columns = inverters_data.columns[inverters_data.columns.str.contains('Irradiance')].to_list()
    for irradiance in irradiance_columns:
        if 'curated' in irradiance:
            irradiance_curated_column = irradiance
        else:
            irradiance_poaavg_column = irradiance

    #print(irradiance_curated_column)

    site_list = list(set(
        [re.search(r'\[\w.+', name).group().replace(']', "").replace('[', "") for name in irradiance_columns]))

    if len(site_list) > 1:
        print('More than one site in irradiance data')
        exit()
    else:
        site = site_list[0]

    budget_pr_df = inverters_data.loc[:, inverters_data.columns.str.contains('Timestamp')]
    budget_pr_df[site + " Budget PR"] = [budget_pr.loc[site,str(datetime.strptime(str(row['Timestamp']), '%Y-%m-%d %H:%M:%S').date().replace(day = 1))] for index,row in budget_pr_df.iterrows()]



    site_info = {}
    site_info['Site'] = site
    site_info['General Info'] = general_info.loc[site, :]
    site_info['Budget PR'] = budget_pr.loc[site, :]
    site_info['Budget PR table'] = budget_pr_df
    site_info['Component Info'] = component_data.loc[component_data['Site'] == site]

    # Get irradiance data and check for empty data
    irradiance_data = inverters_data.loc[:, inverters_data.columns.str.contains('Irradiance|Timestamp')]
    irradiance_data_nafilled = inverters_data.loc[:, inverters_data.columns.str.contains('Irradiance|Timestamp')].fillna('No data')
    days_under_analysis = pd.to_datetime(irradiance_data['Timestamp']).dt.date.drop_duplicates()
    "months_under_analysis = pd.to_datetime(irradiance_data['Timestamp']).dt.month.drop_duplicates()"
    months_under_analysis = pd.to_datetime(irradiance_data['Timestamp']).apply(lambda x: x.strftime('%m-%Y')).drop_duplicates()


    site_info['Days'] = days_under_analysis
    site_info['Months'] = months_under_analysis

    # Active periods aka Irradiance > 20 W/m2
    active_irradiance = irradiance_data.loc[(irradiance_data[irradiance_curated_column] >= 20)]
    active_index = active_irradiance.index

    # To get periods without data
    empty_irradiance = irradiance_data_nafilled.loc[(irradiance_data_nafilled[irradiance_curated_column] == 'No data')]

    all_inverter_power_data_dict = {}

    for inverter in inverter_list:
        inverter_power_data_dict = {}
        capacity = float(site_info['Component Info'].loc[site_info['Component Info']['Component'] == inverter][
                             'Nominal Power DC'].values)

        inverter_power_data_dict['Power Data'] = inverters_data.loc[
            active_index, inverters_data.columns.str.contains(inverter + '|Timestamp')]

        ac_power_column = inverter_power_data_dict['Power Data'].columns[
            inverter_power_data_dict['Power Data'].columns.str.contains('AC')].values[0]
        dc_power_column = inverter_power_data_dict['Power Data'].columns[
            inverter_power_data_dict['Power Data'].columns.str.contains('DC')].values[0]

        inverter_power_data_dict['Power Data']['Efficiency ' + inverter] = (
                inverter_power_data_dict['Power Data'][ac_power_column] / inverter_power_data_dict['Power Data'][
                dc_power_column])
        inverter_power_data_dict['Power Data']['Expected Power ' + inverter] = active_irradiance[
            irradiance_curated_column].multiply(capacity / 1000) * budget_pr_df[site + " Budget PR"]

        inverter_power_data_dict['Power Data']['Ideal Power ' + inverter] = active_irradiance[
            irradiance_curated_column].multiply(capacity / 1000)

        inverter_power_data_dict['Power Data']['Irradiance'] = active_irradiance[irradiance_curated_column]

        all_inverter_power_data_dict[inverter] = inverter_power_data_dict
        power_data = inverter_power_data_dict['Power Data']

        # Create df with all data
        try:
            df_to_add = power_data.drop(columns=['Timestamp', 'Irradiance'])  #, 'Day', 'Month'
            all_inverter_data_df = pd.concat([all_inverter_data_df, df_to_add], axis=1)
        except NameError:
            print('Creating dataframe with all inverters')
            all_inverter_data_df = power_data

    return inverter_list, site_info, all_inverter_power_data_dict,all_inverter_data_df, days_under_analysis, months_under_analysis

def get_percentage_of_timestamp(timestamp, rounded_timestamp, granularity: int = 15):

    difference = abs(rounded_timestamp-timestamp).seconds/60

    percentage_of_timestamp = (granularity - difference)/granularity


    return percentage_of_timestamp


def read_analysis_df_and_correct_date(reportfiletemplate, date, roundto: int = 15):
    day = int(date[-2:])
    month = int(date[-5:-3])
    year = int(date[:4])

    freq = str(roundto) + 'min'
    df_incidents_analysis = pd.read_excel(reportfiletemplate, sheet_name='Analysis of CE', engine="openpyxl")
    df_tracker_analysis = pd.read_excel(reportfiletemplate, sheet_name='Analysis of tracker incidents', engine="openpyxl")

    timestamps = df_incidents_analysis['Time']
    timestamps = pd.Series(timestamps).dt.round(freq, 'shift_backward')
    df_incidents_analysis['Time'] = timestamps
    df_tracker_analysis['Time'] = timestamps

    for timestamp in timestamps:
        newtimestamp = timestamp.replace(year=year, month=month, day=day)
        df_incidents_analysis['Time'] = df_incidents_analysis['Time'].replace(timestamp, newtimestamp)
        df_tracker_analysis['Time'] = df_tracker_analysis['Time'].replace(timestamp, newtimestamp)

    #print(df_tracker_analysis)

    return df_incidents_analysis,df_tracker_analysis

def change_times_to_str(df, active: bool = False):


    df['Event Start Time'] = [str(time) for time in df['Event Start Time']]
    if not active:
        df['Event End Time'] = [str(time) for time in df['Event End Time']]


    return df

def remove_milliseconds(df, end_time: bool = False):
    ''' Removes milliseconds from timestamps'''
    if end_time == True:
        df['Event Start Time'] = [str(timestamp) for timestamp in df['Event Start Time']]
        df['Event End Time'] = [str(timestamp) for timestamp in df['Event End Time']]

        for index, row in df.iterrows():
            stime = df.loc[index, 'Event Start Time']
            etime = df.loc[index, 'Event End Time']
            if "." in stime:
                dot_position = stime.index('.')
                df.loc[index, 'Event Start Time'] = str(stime[:dot_position])
            if "." in etime:
                dot_position = etime.index('.')
                df.loc[index, 'Event End Time'] = str(etime[:dot_position])
    else:
        df['Event Start Time'] = [str(timestamp) for timestamp in df['Event Start Time']]

        for index, row in df.iterrows():
            stime = df.loc[index, 'Event Start Time']
            if "." in stime:
                dot_position = stime.index('.')
                df.loc[index, 'Event Start Time'] = str(stime[:dot_position])


    return df

def remove_incidents_component_type(df, comp_type, column: str = "Related Component"):
    ''' not done, add escape in case comp_type is not string '''
    comp_type = str(comp_type)
    remove_index = []
    component_list = []
    for index, row in df.iterrows():
        rel_comp = df.loc[index,column]
        if comp_type in rel_comp:
            remove_index.append(index)
            component_list.append(rel_comp)

    df_final = df.drop(remove_index)
    df_final = df_final.reset_index(None,drop = True)

    return df_final

def remove_after_sunset_events(site_list, df_input, df_info_sunlight, active_df : bool = False ,tracker: bool = False):
    if tracker == False:
        df_final = df_input
        # This script goes through the different site's dataframes one by one and checks all of the dataframe at once
        for site in site_list:
            # get site dataframe
            df = df_input[site]
            #check if dataframe is not empty, in case it is then go straight to assigning the new df to the new df list
            if not df.empty:
                # Get index of site in info_sunlight dataframe
                index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site].index.values
                index_site = int(index_site_array[0])

                # Get sunrise and sunset hour for a given site on the info_sunlight dataframe
                sunrise = df_info_sunlight.at[index_site, 'Time of operation start']
                if type(sunrise) == str:
                    sunrise = datetime.strptime(sunrise, '%Y-%m-%d %H:%M:%S')

                sunset = df_info_sunlight.at[index_site, 'Time of operation end']
                if type(sunset) == str:
                    sunset = datetime.strptime(sunset, '%Y-%m-%d %H:%M:%S')

                # Create test columns
                start_time_list = df['Event Start Time']
                start_time_list = [datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S') for timestamp in start_time_list]
                df['Event Start Time test'] = start_time_list
                if active_df == False:
                    end_time_list = df['Event End Time']
                    end_time_list = [datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S') for timestamp in end_time_list]
                    df['Event End Time test'] = end_time_list

                # Select all entries in dataframe that have start time before sunset or end time after sunrise a.k.a.
                # the wanted events;
                df_final_site = df.loc[df['Event Start Time test'] < sunset]
                if active_df == False:
                    df_final_site = df_final_site.loc[df['Event End Time test'] > sunrise]

                # get rid of the auxiliary columns
                df_final_site = df_final_site.drop(columns=['Event Start Time test'])
                if active_df == False:
                    df_final_site = df_final_site.drop(columns=['Event End Time test'])

                # add new dataframe to new dataframe list
                df_final[site] = df_final_site

    else:
        # this script goes through the incidents one by one to check which to keep because there are entries for
        # various sites in the same dataframe
        df = df_input
        if not df.empty:
            df['Outside of operation period'] = ""
            for index, row in df.iterrows():
                site_event = df.at[index,"Site Name"]
                index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site_event].index.values
                index_site = int(index_site_array[0])

                sunrise = df_info_sunlight.at[index_site, 'Time of operation start']
                sunset = df_info_sunlight.at[index_site, 'Time of operation end']

                start_time_event = df.at[index, "Event Start Time"]
                if type(start_time_event) == str:
                    start_time_event = datetime.strptime(start_time_event, '%Y-%m-%d %H:%M:%S')
                if start_time_event >= sunset:
                    df.loc[index,'Outside of operation period'] = "x"

                if active_df == False:
                    end_time_event = df.at[index, "Event End Time"]
                    if type(end_time_event) == str:
                         end_time_event = datetime.strptime(end_time_event, '%Y-%m-%d %H:%M:%S')
                    if end_time_event <= sunrise:
                        df.loc[index, 'Outside of operation period'] = "x"


                df_final = df.loc[df['Outside of operation period'] != "x"]
                df_final = df_final.drop(columns=['Outside of operation period'])
                df_final = df_final.reset_index(None, drop=True)
        else:
            df_final = df


    return df_final

def rounddatesactive_15m(site, df, freq: int = 15):

    ''' Rounds date of active events (start time) to nearest 15m timestamp

    Returns df'''
    round_to = str(freq) + 'min'

    try:
        to_round_startdate = df['Event Start Time']
        to_round_startdate = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp in to_round_startdate]

        rounded_startdate = pd.Series(to_round_startdate).dt.round(round_to, 'shift_backward')

        rounded_startdate.index = df.index.to_list()

        df['Rounded Event Start Time'] = rounded_startdate

    except AttributeError:
        print('No new active events on this day for ' + site)

    return df

def rounddatesclosed_15m(site, df, freq: int = 15):
    '''Rounds date of closed events (start time and end time) to nearest 15m timestamp

    Retunrs df '''
    round_to = str(freq) + 'min'

    try:
        to_round_startdate = df['Event Start Time']
        to_round_startdate = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp in to_round_startdate]

        to_round_enddate = df['Event End Time']
        to_round_enddate = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp in to_round_enddate]

        rounded_startdate = pd.Series(to_round_startdate).dt.round(round_to, 'shift_backward')
        rounded_enddate = pd.Series(to_round_enddate).dt.round(round_to, 'shift_forward')

        rounded_startdate.index = df.index.to_list()
        rounded_enddate.index = df.index.to_list()

        df['Rounded Event Start Time'] = rounded_startdate
        df['Rounded Event End Time'] = rounded_enddate

    except AttributeError:
        print('No new closed events on this day for ' + site)

    return df

def verify_read_time_of_operation(site, day, stime, etime):
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.
    layout = [[sg.Text('Are these hours correct on the ' + str(day) + ' on ' + site + '?')],
              [sg.HorizontalSeparator(pad=((10, 10), (2, 10)))],
              [sg.Text('Sunrise hour: '), sg.Text(str(stime))],
              [sg.Text('Sunset hour: '), sg.Text(str(etime))],
              [sg.Button('Submit'), sg.Button('Change hours'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Daily Monitoring Report', layout)
    # Event Loop to process "events" and get the "values" of the inputs

    while True:
        event, values = window.read()

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            break
        if event == 'Submit':
            print('Submitted for ' + site + ':')
            print(stime)
            print(etime)
            break
        if event == 'Change hours':
            stime, etime = input_time_operation_site(site, str(day))
            break
    window.close()

    return stime, etime





# Xlsxwriter related custom functions ----------------------------------------------------------

def get_rowindex_and_columnletter(cell):
    cell_letter_code = re.search(r'([\w]+)([\d]+)', cell)
    rowindex = cell_letter_code.group(0)
    column_letter = cell_letter_code.group(1)
    return rowindex, column_letter

def get_col_widths(dataframe):
    # First we find the maximum length of the index column
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) for col in dataframe.columns]

def range_write_blank(worksheet, col_start, col_end, row_start, row_end, cel_format):
    '''
    Writes blank to range of cells cel_format
    Takes:
    worksheet as Worksheet object
    col_start as uppercase letter string
    col_end as uppercase letter string
    row_start as int
    row_end as int
    cel_format as XlsXWriter Format object
    '''
    letters = list(string.ascii_uppercase)
    for r in range(row_start, row_end):
        for c in range(letters.index(col_start), letters.index(col_end)+1):
            worksheet.write_blank(r, c, '', cel_format)



#File creation/edit/removal-------------------------------------------------------------------------------------

def add_incidents_to_excel(dest_file,site_list,df_list_active, df_list_closed,df_info_sunlight, final_irradiance_data):

    '''USAGE: add_incidents_to_excel(destiny_file,site_list,df_list_active,df_list_closed)'''

    append_df_to_excel(dest_file, df_info_sunlight, sheet_name='Info', startrow = 0)
    append_df_to_excel(dest_file, final_irradiance_data, sheet_name='Irradiance', startrow=0)

    for site in site_list:
        if "LSBP - " in site or "LSBP – " in site:
            onlysite = site[7:]
        else:
            onlysite = site
        if onlysite[-1:] == " ":
            active_sheet_name = onlysite + 'Active'
            closed_sheet_name = onlysite[:len(onlysite)-1]
        else:
            active_sheet_name = onlysite + ' Active'
            closed_sheet_name = onlysite



        df_active= df_list_active[site]
        df_closed= df_list_closed[site]


        df_closed['Status of incident'] = 'Closed'
        df_active['Status of incident'] = 'Active'
        df_active['Action required'] = ''

        append_df_to_excel(dest_file, df_closed, sheet_name= closed_sheet_name)

        print('Active events of ' + site + ' added')
        append_df_to_excel(dest_file, df_active, sheet_name= active_sheet_name)
        print('Closed events of ' + site + ' added')


    return

def add_tracker_incidents_to_excel(dest_tracker_file, df_tracker_active, df_tracker_closed, df_tracker_info):

    '''USAGE: add_tracker_incidents_to_excel(dest_file, df_tracker_active, df_tracker_closed, df_tracker_info)'''

    append_df_to_excel(dest_tracker_file, df_tracker_info, sheet_name='Trackers info', startrow = 0)
    print('Tracker Info added')

    df_tracker_closed['Status of incident'] = 'Closed'
    df_tracker_active['Status of incident'] = 'Active'
    df_tracker_active['Action required'] = ''

    append_df_to_excel(dest_tracker_file, df_tracker_active, sheet_name='Active tracker incidents')
    print('Active tracker incidents added')
    append_df_to_excel(dest_tracker_file, df_tracker_closed, sheet_name='Closed tracker incidents')
    print('Closed tracker incidents added')

    return

def add_events_to_final_report(reportfile, df_list_active, df_list_closed,df_tracker_active, df_tracker_closed):

    final_active_events_list = pd.concat(list(df_list_active.values()))
    final_closed_events_list = pd.concat(list(df_list_closed.values()))


    if not final_active_events_list.empty:
        append_df_to_excel(reportfile, final_active_events_list, sheet_name='Active Events', startrow=0)
        print('Active events added')
    else:
        print('No active events to be added')

    if not final_closed_events_list.empty:
        append_df_to_excel(reportfile, final_closed_events_list, sheet_name='Closed Events', startrow=0)
        print('Closed events added')
    else:
        print('No closed events to be added')

    if not df_tracker_active.empty:
        append_df_to_excel(reportfile, df_tracker_active, sheet_name='Active tracker incidents', startrow=0)
        print('Tracker active events added')
    else:
        print('No tracker active events to be added')

    if not df_tracker_closed.empty:
        append_df_to_excel(reportfile, df_tracker_closed, sheet_name='Closed tracker incidents', startrow=0)
        print('Tracker closed events added')
    else:
        print('No tracker closed events to be added')

    return

def add_analysis_to_reportfile(reportfile,df_incidents_analysis,df_tracker_analysis, df_info_sunlight):


    # Add Info sheet
    append_df_to_excel(reportfile, df_info_sunlight, sheet_name='Info', startrow=0)

    # Add component failure analysis
    append_df_to_excel(reportfile, df_incidents_analysis, sheet_name='Analysis of CE', startrow=0)

    # Add component failure analysis
    append_df_to_excel(reportfile, df_tracker_analysis, sheet_name='Analysis of tracker incidents', startrow=0)


    return

def update_dump_file(irradiance_files,all_irradiance_file, data_type:str = 'Irradiance'):
    df_all_irradiance = pd.read_excel(all_irradiance_file, engine='openpyxl')

    df_irradiance_day_list = [pd.read_excel(file, engine='openpyxl') for file in irradiance_files]
    df_all_irradiance_list = df_irradiance_day_list.append(df_all_irradiance)

    df_all_irradiance_new = pd.concat(df_irradiance_day_list)
    df_all_irradiance_new['Timestamp'] = [datetime.strptime(str(timestamp), '%Y-%m-%d %H:%M:%S') for timestamp in df_all_irradiance_new['Timestamp']]
    df_all_irradiance_new = df_all_irradiance_new.loc[:,
                            ~df_all_irradiance_new.columns.str.contains('^Unnamed')].drop_duplicates(
        subset=['Timestamp'], keep='first', ignore_index=True).sort_values(by=['Timestamp'], ascending=[True],
                                                                           ignore_index=True)

    writer_irr = pd.ExcelWriter(all_irradiance_file, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_numbers': True}})
    workbook_irr = writer_irr.book

    df_all_irradiance_new.to_excel(writer_irr, sheet_name='All ' + str(data_type) , index=False)

    writer_irr.save()


    return df_all_irradiance_new

"""def add_events_to_final_reportv2(reportfile, df_list_active, df_list_closed,df_tracker_active, df_tracker_closed):
    x = 0
    for site in df_list_active.keys():
        df = df_list_active[site]
        if not df.empty:
            if x < 1:
                append_df_to_excel(reportfile, df_list_active[site], sheet_name='Active Events', startrow=0)
                print(site + ' active events added')
                x+=1
            else:
                append_df_to_excel_existing(reportfile,df_list_active[site], sheet_name= 'Active Events')
                print(site + ' active events added')
    x = 0
    for site in df_list_closed.keys():
        df = df_list_closed[site]
        if not df.empty:
            if x < 1:
                append_df_to_excel(reportfile, df_list_closed[site], sheet_name='Closed Events', startrow = 0)
                print(site + ' closed events added')
                x+=1
            else:
                append_df_to_excel_existing(reportfile, df_list_closed[site], sheet_name='Closed Events')
                print(site + ' closed events added')

    if not df_tracker_active.empty:
        append_df_to_excel(reportfile, df_tracker_active, sheet_name='Active tracker incidents', startrow=0)
        print('Tracker active events added')
    else:
        print('No tracker active events to be added')

    if not df_tracker_closed.empty:
        append_df_to_excel(reportfile, df_tracker_closed, sheet_name='Closed tracker incidents', startrow=0)
        print('Tracker closed events added')
    else:
        print('No tracker closed events to be added')

    return


def add_analysis_to_reportfilev2(reportfile,df_incidents_analysis,df_tracker_analysis, df_info_sunlight):


    # Add Info sheet
    append_df_to_excel(reportfile, df_info_sunlight, sheet_name='Info', startrow=0)

    # Add component failure analysis
    append_df_to_excel(reportfile, df_incidents_analysis, sheet_name='Analysis of CE', startrow=0)

    # Add component failure analysis
    append_df_to_excel(reportfile, df_tracker_analysis, sheet_name='Analysis of tracker incidents', startrow=0)


    return""" #alternative versions of add_events/analysis to final report


#Transversal functions -------------------------------------------------------------------------------------------

def rename_dict_keys(d, keys):
    return dict([(keys.get(k), v) for k, v in d.items()])



#Data analysis functions -------------------------------------------------------------------------------------------

def calculate_daily_raw_pr(inverter_data, days_under_analysis,inverter):
    '''From Inverter data (Power AC and Expected Power) calculates Raw PR
    Also uses irradiance to complete Dataframe'''

    ac_power_column = inverter_data.columns[inverter_data.columns.str.contains('AC')].values[0]
    expected_power_column = inverter_data.columns[inverter_data.columns.str.contains('Expected')].values[0]
    ideal_power_column = inverter_data.columns[inverter_data.columns.str.contains('Ideal')].values[0]
    irradiance_column = inverter_data.columns[inverter_data.columns.str.contains('Irradiance')].values[0]

    daily_pr_inverter = {}
    for day in days_under_analysis:
        data_day = inverter_data.loc[inverter_data['Day'] == day]
        actual_power_day = data_day[ac_power_column].sum() / 4
        expected_power_day = data_day[expected_power_column].sum() / 4
        ideal_power_day = data_day[ideal_power_column].sum() / 4

        irradiance_day = data_day[irradiance_column].sum() / 4

        pr_day = actual_power_day / ideal_power_day
        daily_pr_inverter[day] = (pr_day, irradiance_day)

        # print(day, ": ", actual_power_day, " / " , expected_power_day, " / ", pr_day, " / ")

    daily_pr_df = pd.DataFrame.from_dict(daily_pr_inverter, orient='index', columns=[str(inverter) + ' PR %', irradiance_column])
    # print(df)



    return daily_pr_df,irradiance_column

def calculate_daily_corrected_pr(inverter_data, days_under_analysis, inverter, maxexport_capacity_ac):
    '''From Inverter data (Power AC, Expected Power and Max export capacity ) calculates Corrected PR
    Corrected PR, in this case, is the correction for max export capacity.
        Also uses irradiance to complete Dataframe'''

    ac_power_column = inverter_data.columns[inverter_data.columns.str.contains('AC')].values[0]
    expected_power_column = inverter_data.columns[inverter_data.columns.str.contains('Expected')].values[0]
    ideal_power_column = inverter_data.columns[inverter_data.columns.str.contains('Ideal')].values[0]
    irradiance_column = inverter_data.columns[inverter_data.columns.str.contains('Irradiance')].values[0]


    corrected_power_data = inverter_data
    corrected_power_data[expected_power_column] = [maxexport_capacity_ac if power > maxexport_capacity_ac else power for power in
                                                   corrected_power_data[expected_power_column]]
    corrected_power_data[ideal_power_column] = [maxexport_capacity_ac if power > maxexport_capacity_ac else power for
                                                   power in
                                                   corrected_power_data[ideal_power_column]]
    corrected_daily_pr_dict = {}

    for day in days_under_analysis:
        corrected_data_day = corrected_power_data.loc[corrected_power_data['Day'] == day]
        corrected_actual_power_day = corrected_data_day[ac_power_column].sum() / 4
        corrected_expected_power_day = corrected_data_day[expected_power_column].sum() / 4
        corrected_ideal_power_day = corrected_data_day[ideal_power_column].sum() / 4

        irradiance_day = corrected_data_day[irradiance_column].sum() / 4

        corrected_pr_day = corrected_actual_power_day / corrected_ideal_power_day
        corrected_daily_pr_dict[day] = (corrected_pr_day, irradiance_day)

    corrected_daily_pr_df = pd.DataFrame.from_dict(corrected_daily_pr_dict, orient='index',
                                                    columns=[str(inverter) + ' Corrected PR %',
                                                             irradiance_column])
    # print(corrected_daily_pr_df)



    return corrected_daily_pr_df,irradiance_column

def calculate_daily_corrected_pr_focusDC(inverter_data, days_under_analysis, inverter, maxexport_capacity_ac):
    '''From Inverter data (Power AC, Expected Power and Max export capacity ) calculates Corrected PR
    Corrected PR, in this case, is the correction for inverter failures (focus on DC side) and with max export capacity in place
        Also uses irradiance to complete Dataframe'''

    ac_power_column = inverter_data.columns[inverter_data.columns.str.contains('AC')].values[0]
    expected_power_column = inverter_data.columns[inverter_data.columns.str.contains('Expected')].values[0]
    ideal_power_column = inverter_data.columns[inverter_data.columns.str.contains('Ideal')].values[0]
    irradiance_column = inverter_data.columns[inverter_data.columns.str.contains('Irradiance')].values[0]


    corrected_power_data = inverter_data.loc[inverter_data[ac_power_column] > 0]
    corrected_power_data[expected_power_column] = [maxexport_capacity_ac if power > maxexport_capacity_ac else power for power in
                                                   corrected_power_data[expected_power_column]]
    corrected_power_data[ideal_power_column] = [maxexport_capacity_ac if power > maxexport_capacity_ac else power for
                                                   power in
                                                   corrected_power_data[ideal_power_column]]
    corrected_daily_pr_dict = {}

    for day in days_under_analysis:
        corrected_data_day = corrected_power_data.loc[corrected_power_data['Day'] == day]
        corrected_actual_power_day = corrected_data_day[ac_power_column].sum() / 4
        corrected_expected_power_day = corrected_data_day[expected_power_column].sum() / 4
        corrected_ideal_power_day = corrected_data_day[ideal_power_column].sum() / 4

        irradiance_day = corrected_data_day[irradiance_column].sum() / 4

        corrected_pr_day = corrected_actual_power_day / corrected_ideal_power_day
        corrected_daily_pr_dict[day] = (corrected_pr_day, irradiance_day)

    corrected_df = pd.DataFrame.from_dict(corrected_daily_pr_dict, orient='index',
                                                    columns=[str(inverter) + ' - DC focus - Corrected PR %',
                                                             irradiance_column])
    # print(df)


    return corrected_df,irradiance_column

def calculate_monthly_raw_pr(inverter_data, months_under_analysis,inverter):
    '''From Inverter data (Power AC, Expected Power and Max export capacity ) calculates Corrected PR
        Corrected PR, in this case, is the correction for inverter failures (focus on DC side) and with max export capacity in place
        Also uses irradiance to complete Dataframe'''

    ac_power_column = inverter_data.columns[inverter_data.columns.str.contains('AC')].values[0]
    expected_power_column = inverter_data.columns[inverter_data.columns.str.contains('Expected')].values[0]
    ideal_power_column = inverter_data.columns[inverter_data.columns.str.contains('Ideal')].values[0]
    irradiance_column = inverter_data.columns[inverter_data.columns.str.contains('Irradiance')].values[0]

    raw_monthly_pr_dict = {}
    raw_powers_dict_forsite = {}

    for month in months_under_analysis:
        raw_data_month = inverter_data.loc[inverter_data['Month'] == month]
        raw_actual_power_month = raw_data_month[ac_power_column].sum() / 4
        raw_expected_power_month = raw_data_month[expected_power_column].sum() / 4
        raw_ideal_power_month = raw_data_month[ideal_power_column].sum() / 4

        irradiance_month = raw_data_month[irradiance_column].sum() / 4

        raw_pr_month = raw_actual_power_month / raw_ideal_power_month
        raw_monthly_pr_dict[month] = (raw_pr_month, irradiance_month)
        raw_powers_dict_forsite[month] = (raw_actual_power_month, raw_expected_power_month, raw_ideal_power_month)

    raw_monthly_pr_df = pd.DataFrame.from_dict(raw_monthly_pr_dict, orient='index', columns=[
        str(inverter) + ' Raw Monthly PR %', irradiance_column])
    raw_monthly_production_df = pd.DataFrame.from_dict(raw_powers_dict_forsite,
                                                             orient='index',
                                                             columns=[ac_power_column, expected_power_column, ideal_power_column])
    # print(df)

    return raw_monthly_pr_df, raw_monthly_production_df,irradiance_column

def calculate_monthly_corrected_pr_and_production_focusDC(inverter_data, months_under_analysis, inverter, maxexport_capacity_ac):
    '''From Inverter data (Power AC, Expected Power and Max export capacity ) calculates Corrected PR
        Corrected PR, in this case, is the correction for inverter failures (focus on DC side) and with max export capacity in place
        Also uses irradiance to complete Dataframe'''

    ac_power_column = inverter_data.columns[inverter_data.columns.str.contains('AC')].values[0]
    expected_power_column = inverter_data.columns[inverter_data.columns.str.contains('Expected')].values[0]
    ideal_power_column = inverter_data.columns[inverter_data.columns.str.contains('Ideal')].values[0]
    irradiance_column = inverter_data.columns[inverter_data.columns.str.contains('Irradiance')].values[0]

    corrected_power_data = inverter_data.loc[inverter_data[ac_power_column] > 0]
    corrected_power_data[expected_power_column] = [maxexport_capacity_ac if power > maxexport_capacity_ac else power for power in
                                                   corrected_power_data[expected_power_column]]
    corrected_power_data[ideal_power_column] = [maxexport_capacity_ac if power > maxexport_capacity_ac else power for
                                                   power in
                                                   corrected_power_data[ideal_power_column]]

    corrected_monthly_pr_dict = {}
    corrected_powers_dict_forsite = {}

    for month in months_under_analysis:
        corrected_data_month = corrected_power_data.loc[corrected_power_data['Month'] == month]
        corrected_actual_power_month = corrected_data_month[ac_power_column].sum() / 4
        corrected_expected_power_month = corrected_data_month[expected_power_column].sum() / 4
        corrected_ideal_power_month = corrected_data_month[ideal_power_column].sum() / 4

        irradiance_month = corrected_data_month[irradiance_column].sum() / 4

        corrected_pr_month = corrected_actual_power_month / corrected_ideal_power_month
        corrected_monthly_pr_dict[month] = (corrected_pr_month, irradiance_month)
        corrected_powers_dict_forsite[month] = (corrected_actual_power_month, corrected_expected_power_month, corrected_ideal_power_month)

    corrected_monthly_pr_df = pd.DataFrame.from_dict(corrected_monthly_pr_dict, orient='index', columns=[
        str(inverter) + ' Corrected (w/clipping) Monthly PR %', irradiance_column])
    corrected_monthly_production_df = pd.DataFrame.from_dict(corrected_powers_dict_forsite,
                                                                       orient='index',
                                                                       columns=[ac_power_column, expected_power_column, ideal_power_column])
    # print(df)


    return corrected_monthly_pr_df, corrected_monthly_production_df,irradiance_column

def calculate_monthly_corrected_pr_and_production(inverter_data, months_under_analysis,inverter, capacity_ac):
    '''From Inverter data (Power AC, Expected Power and Max export capacity ) calculates Corrected PR
        Corrected PR, in this case, is the correction for inverter failures (focus on DC side) and with max export capacity in place
        Also uses irradiance to complete Dataframe'''

    ac_power_column = inverter_data.columns[inverter_data.columns.str.contains('AC')].values[0]
    expected_power_column = inverter_data.columns[inverter_data.columns.str.contains('Expected')].values[0]
    ideal_power_column = inverter_data.columns[inverter_data.columns.str.contains('Ideal')].values[0]
    irradiance_column = inverter_data.columns[inverter_data.columns.str.contains('Irradiance')].values[0]

    corrected_power_data = inverter_data
    corrected_power_data[expected_power_column] = [capacity_ac if power > capacity_ac else power for power in
                                                   corrected_power_data[expected_power_column]]
    corrected_power_data[ideal_power_column] = [capacity_ac if power > capacity_ac else power for power in
                                                   corrected_power_data[ideal_power_column]]

    corrected_monthly_pr_dict = {}
    corrected_powers_dict_forsite = {}

    for month in months_under_analysis:
        corrected_data_month = corrected_power_data.loc[corrected_power_data['Month'] == month]
        corrected_actual_power_month = corrected_data_month[ac_power_column].sum() / 4
        corrected_expected_power_month = corrected_data_month[expected_power_column].sum() / 4
        corrected_ideal_power_month = corrected_data_month[ideal_power_column].sum() / 4

        irradiance_month = corrected_data_month[irradiance_column].sum() / 4

        corrected_pr_month = corrected_actual_power_month / corrected_ideal_power_month
        corrected_monthly_pr_dict[month] = (corrected_pr_month, irradiance_month)
        corrected_powers_dict_forsite[month] = (corrected_actual_power_month, corrected_expected_power_month, corrected_ideal_power_month)

    corrected_monthly_pr_df = pd.DataFrame.from_dict(corrected_monthly_pr_dict, orient='index', columns=[
        str(inverter) + ' Corrected (w/clipping) Monthly PR %', irradiance_column])
    corrected_monthly_production_df = pd.DataFrame.from_dict(corrected_powers_dict_forsite,
                                                             orient='index',
                                                             columns=[ac_power_column, expected_power_column, ideal_power_column])
    # print(df)

    return corrected_monthly_pr_df, corrected_monthly_production_df,irradiance_column

def calculate_expected_energy(site,start_timestamp, end_timestamp, budget_export, budget_irradiance, actual_irradiance_site):
    # Calculate Expected Energy in period
    if start_timestamp.month == end_timestamp.month and start_timestamp.year == end_timestamp.year:
        expected_energy_info = {}
        seconds_in_period = (end_timestamp - start_timestamp).total_seconds()
        seconds_in_month = calendar.monthrange(start_timestamp.year, start_timestamp.month)[1] * 24 * 3600
        percentage_of_month = seconds_in_period / seconds_in_month

        actual_irradiance_slice = actual_irradiance_site.loc[
            (actual_irradiance_site.index <= end_timestamp) & (actual_irradiance_site.index >= start_timestamp)]

        budget_energy_month = budget_export.loc[site, str(start_timestamp.replace(day=1).date())]
        budget_irradiance_month = budget_irradiance.loc[site, str(start_timestamp.replace(day=1).date())]
        budget_energy_slice = percentage_of_month * budget_energy_month
        budget_irradiance_slice = percentage_of_month * budget_irradiance_month

        expected_energy_slice = (budget_energy_slice * (actual_irradiance_slice.sum() / 4)) / budget_irradiance_slice

        expected_energy = (budget_export.loc[site, str(start_timestamp.replace(day=1).date())] * (
                actual_irradiance_site.sum() / 4)) / budget_irradiance.loc[
                              site, str(start_timestamp.replace(day=1).date())]

        expected_energy_info[str(start_timestamp.replace(day=1).date())] = {
            "Budget Irradiance Month": budget_irradiance_month,
            "Budget Export Month": budget_energy_month,
            "Percentage of month": percentage_of_month,
            "Budget Irradiance Period": budget_irradiance_slice,
            "Budget Export Period": budget_energy_slice,
            "Actual Irradiance Period": actual_irradiance_slice.sum() / 4000,
            "Expected Energy Period": expected_energy}

    else:
        # If not restricted to a month, the script will separate by monhts and calculate the expected energy in each
        # month's slice, in the end it will sum it all to give an expected energy for the period in analysis
        date_range = pd.date_range(start_timestamp.replace(day=1), end_timestamp.replace(day=1),
                                   freq=pd.offsets.MonthBegin(1))
        expected_energy_in_period = {}
        expected_energy_info = {}
        for date in date_range:
            budget_energy_month = budget_export.loc[site, str(date.date())]
            budget_irradiance_month = budget_irradiance.loc[site, str(date.date())]
            month = date.month

            if month == start_timestamp.month:
                seconds_in_period_month = (pd.Timestamp(start_timestamp.year, start_timestamp.month,
                                                        calendar.monthrange(start_timestamp.year,
                                                                            start_timestamp.month)[1], 23, 59,
                                                        59) - start_timestamp).total_seconds()
                percentage_of_month = seconds_in_period_month / (
                        calendar.monthrange(start_timestamp.year, month)[1] * 24 * 3600)
                actual_irradiance_slice = actual_irradiance_site.loc[(actual_irradiance_site.index <= pd.Timestamp(
                    start_timestamp.year, start_timestamp.month,
                    calendar.monthrange(start_timestamp.year, start_timestamp.month)[1], 23, 59, 59)) & (
                                                                             actual_irradiance_site.index >= start_timestamp)]


            elif month == end_timestamp.month:
                seconds_in_period_month = (
                        end_timestamp - pd.Timestamp(end_timestamp.year, end_timestamp.month, 1)).total_seconds()
                percentage_of_month = seconds_in_period_month / (
                        calendar.monthrange(end_timestamp.year, month)[1] * 24 * 3600)
                actual_irradiance_slice = actual_irradiance_site.loc[(actual_irradiance_site.index <= end_timestamp) & (
                        actual_irradiance_site.index >= pd.Timestamp(start_timestamp.year, end_timestamp.month, 1))]


            else:
                percentage_of_month = 1

                actual_irradiance_slice = actual_irradiance_site.loc[(actual_irradiance_site.index <= pd.Timestamp(
                        date.year, month,
                        calendar.monthrange(date.year, month)[1], 23, 59, 59)) & (
                                                                             actual_irradiance_site.index >= pd.Timestamp(date.year, month,1, 0, 0, 0))]


            budget_energy_slice = percentage_of_month * budget_energy_month
            budget_irradiance_slice = percentage_of_month * budget_irradiance_month

            if not percentage_of_month == 0:
                expected_energy_slice = (budget_energy_slice * (actual_irradiance_slice.sum() / 4)) / budget_irradiance_slice
            else:
                expected_energy_slice = 0


            expected_energy_in_period[str(date.date())] = expected_energy_slice


            expected_energy_info[str(date.date())] = {"Budget Irradiance Month": budget_irradiance_month,
                                                      "Budget Export Month": budget_energy_month,
                                                      "Percentage of month": percentage_of_month,
                                                      "Budget Irradiance Period": budget_irradiance_slice,
                                                      "Budget Export Period": budget_energy_slice,
                                                      "Actual Irradiance Period": actual_irradiance_slice.sum() / 4000,
                                                      "Expected Energy Period": expected_energy_slice}


        expected_energy = sum(expected_energy_in_period.values())



    return expected_energy, expected_energy_info




def get_all_units_from_operation_hours(df_operation_hours):

    inverters = df_operation_hours.columns.drop('Timestamp')
    inverter_operation = {}

    for inverter in inverters:
        print(inverter)
        df_inverter = df_operation_hours[['Timestamp', inverter]].dropna().reset_index(None, drop=True)
        df_inverter_first_hour = df_inverter.loc[df_inverter[inverter] == 1]['Timestamp'][0]
        df_inverter_last_hour = df_inverter['Timestamp'][len(df_inverter) - 1]

        list_1 = list(df_inverter[inverter])
        list_2 = list(df_inverter[inverter])
        list_1.pop(0)
        list_2.pop(len(list_2) - 1)

        df_inverter['Diff'] = list(np.subtract(np.array(list_1), np.array(list_2))).insert(len(list_2), 0)
        df_change = df_inverter.loc[df_inverter['Diff'] < 0]
        first_hour_timestamp = df_inverter.loc[df_inverter[inverter] == 1]

        if len(df_change) == 0:
            inverter_operation[inverter] = [df_inverter_first_hour, df_inverter_last_hour]

        elif len(df_change) == 1:
            inverter_operation[inverter] = [df_inverter_first_hour, list(df_change['Timestamp'])[0]]

            inverter_name = inverter + ".r" + str(2)
            inverter_operation[inverter_name] = [list(df_change['Timestamp'])[0], df_inverter_last_hour]

        else:
            n_changes = len(df_change)
            print("Found " + str(n_changes) + " changes on ", inverter)

            for i in range(n_changes):
                # print(i)
                if i == 0:
                    inverter_operation[inverter] = [df_inverter_first_hour, list(df_change['Timestamp'])[i]]

                elif i == n_changes:
                    inverter_name = inverter + ".r" + str(i + 1)
                    inverter_operation[inverter_name] = [list(df_change['Timestamp'])[i], df_inverter_last_hour]
                else:
                    inverter_name = inverter + ".r" + str(i + 1)
                    inverter_operation[inverter_name] = [list(df_change['Timestamp'])[i],
                                                         list(df_change['Timestamp'])[i + 1]]



    return inverter_operation

def complete_dataset_inverterops_data(incidents_site, inverter_operation,df_operation_hours):
    for index, row in incidents_site.iterrows():
        component = row['Related Component']
        incident_time = row['Event Start Time']
        #print(component, incident_time)

        # Type of component
        if "Block" in component:
            #print('here')
            incidents_site.loc[index, 'Component Type'] = "Inverter Block"
            incidents_site.loc[index, 'Unit Component'] = "N/A"
            incidents_site.loc[index, 'Operation Time'] = "N/A"

        elif "LSBP" in component:
            incidents_site.loc[index, 'Component Type'] = "Site"
            incidents_site.loc[index, 'Unit Component'] = "N/A"
            incidents_site.loc[index, 'Operation Time'] = "N/A"

        elif "CB" in component or "String" in component:
            incidents_site.loc[index, 'Component Type'] = "Combiner Box"
            incidents_site.loc[index, 'Unit Component'] = "N/A"
            incidents_site.loc[index, 'Operation Time'] = "N/A"

        else:
            incidents_site.loc[index, 'Component Type'] = "Inverter"

            component_number = re.search(r'\d.*', component).group()

            inverter_operation_time_info = [inverter for inverter in inverter_operation if
                                            str(component_number) in inverter]

            for unit in inverter_operation_time_info:
                stime = inverter_operation[unit][0]
                etime = inverter_operation[unit][1]

                if incident_time > stime and incident_time < etime:
                    incidents_site.loc[index, 'Unit Component'] = unit
                    rounded_incident_time = incident_time.round('15min', 'shift_backward')

                    # print(incident_time, rounded_incident_time)

                    incident_operation_time = \
                    df_operation_hours.loc[df_operation_hours['Timestamp'] == rounded_incident_time][component].values[
                        0]

                    changed = False
                    while np.isnan(incident_operation_time):
                        rounded_incident_time = rounded_incident_time - pd.Timedelta(minutes=15)
                        incident_operation_time = \
                        df_operation_hours.loc[df_operation_hours['Timestamp'] == rounded_incident_time][
                            component].values[0]
                        changed = True
                    if changed == True:
                        print("Changed rounded time to forward timestamp because backward was NaN, new timestamp: ",
                              rounded_incident_time)

                    incidents_site.loc[index, 'Operation Time'] = float(incident_operation_time)
                else:
                    continue



    return incidents_site

def complete_dataset_capacity_data(df_list, all_component_data):
    for site in df_list.keys():
        incidents_site = df_list[site]
        print(type(incidents_site))
        if not type(incidents_site) == str:
            for index, row in incidents_site.iterrows():
                site = row['Site Name']
                component = row['Related Component']

                try:
                    capacity = all_component_data.loc[(all_component_data['Site'] == site)
                                                  & (all_component_data['Component'] == component)]["Nominal Power DC"].values[0]
                except IndexError:
                    capacity = "NA"

                # Add capacity
                incidents_site.loc[index, 'Capacity Related Component'] = capacity

            df_list[site] = incidents_site

    return df_list

def complete_dataset_existing_incidents(df_list, df_dmr):
    for site in df_list.keys():
        print("Completing dataset on " + site)
        incidents_site = df_list[site]
        df_dmr_site = df_dmr.loc[df_dmr['Site Name'] == site]
        if type(df_dmr_site) == str:
            print("No previous active events")

        elif type(incidents_site) == str:
            print("No active events, adding previously active events")
            incidents_site = df_dmr_site

        else:
            incidents_site = pd.concat([incidents_site,df_dmr_site])

        df_list[site] = incidents_site

    return df_list



def timeframe_of_analysis_with_opshours(df_operation_hours):

    start_date_data = df_operation_hours['Timestamp'][0].date()
    end_date_data = df_operation_hours['Timestamp'][len(df_operation_hours) - 1].date()

    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.

    layout = [[sg.Text('Choose number of points:', pad=((2, 10), (2, 5)))],
              [sg.Radio('10', group_id="datapoints", default=False, key="-10DP-"),
               sg.Radio('100', group_id="datapoints", default=True, key="-100DP-"),
               sg.Radio('500', group_id="datapoints", default=False, key="-500DP-")],
              [sg.Text('Enter date of start of analysis', pad=((2, 10), (2, 5)))],
              [sg.CalendarButton('Choose start date', target='-SCAL-',
                                 default_date_m_d_y=(start_date_data.month, start_date_data.day, start_date_data.year),
                                 format="%Y-%m-%d"),
               sg.In(default_text=str(start_date_data), key='-SCAL-', text_color='black', size=(16, 1),
                     enable_events=True, readonly=True, visible=True)],
              [sg.CalendarButton('Choose end date', target='-ECAL-',
                                 default_date_m_d_y=(end_date_data.month, end_date_data.day, end_date_data.year),
                                 format="%Y-%m-%d"),
               sg.In(default_text=str(end_date_data), key='-ECAL-', text_color='black', size=(16, 1),
                     enable_events=True, readonly=True, visible=True)],
              [sg.Button('Submit'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Choose timeframe of analysis', layout)

    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read(timeout=100)

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            window.close()
            return None,None,None
            break

        if event == 'Submit':

            sdate = values['-SCAL-']
            edate = values['-ECAL-']

            if len(sdate) == 0:
                sdate = str(start_date_data)
            #print(values.keys())
            for key in values.keys():
                #print(key)
                if values[key] == True:
                    datapoints = re.search(r'\d+', key).group()
                    #print(datapoints)

            window.close()
            return sdate, edate, datapoints

    window.close()


    return






# Analysis of incidents ---------------------------------------------------------------------------------------------
def analysis_closed_incidents(site, index_site, df_incidents_analysis,df_closed_events, df_info_sunlight):


    max_percentage = "{:.2%}".format(1)
    min_percentage = "{:.2%}".format(0)
    capacity_site = df_info_sunlight.at[index_site, 'Capacity']

    starttime_day = df_info_sunlight.at[index_site, 'Time of operation start']
    endtime_day = df_info_sunlight.at[index_site, 'Time of operation end']

    if not df_closed_events.empty:
        try:
            mintimestamp = df_closed_events['Rounded Event Start Time'].min()
            maxtimestamp = df_closed_events['Rounded Event End Time'].max()

            if mintimestamp < starttime_day:
                mintimestamp = starttime_day
            if maxtimestamp > endtime_day:
                maxtimestamp = endtime_day


        except KeyError:
            print('KeyError: There were no new closed events in this day on ' + site)

        try:
            index_mint = df_incidents_analysis[
                df_incidents_analysis['Time'] == mintimestamp].index.values  # gets starting time row index
            int_index_mint = index_mint[0]  # turns index from numpy.ndarray to integer

            index_maxt = df_incidents_analysis[
                df_incidents_analysis['Time'] == maxtimestamp].index.values  # gets ending time row index
            int_index_maxt = index_maxt[0]  # turns index

            for index in range(int_index_mint, int_index_maxt):
                sum = 0
                for index_not, row in df_closed_events.iterrows():
                    if "roducing" in row['Component Status'] and not row['Curtailment Event'] == "x":
                        if df_incidents_analysis.loc[index, 'Time'] <= row['Rounded Event End Time'] and \
                            df_incidents_analysis.loc[index, 'Time'] >= row['Rounded Event Start Time']:
                            sum += row['Capacity Related Component']
                    else:
                        continue

                    percentage = "{:.2%}".format(sum / capacity_site)
                    if float(percentage[:-1]) > float(max_percentage[:-1]):
                        percentage_final = min_percentage    #test1 - max_percentage
                    else:
                        percentage_final = "{:.2%}".format(1-(sum / capacity_site))
                    df_incidents_analysis.loc[index, site] = percentage_final   #test1 max_percentage -


        except KeyError:
            print('KeyError: There were no approved closed events in this day on ' + site)
        except NameError:
            print('NameError: There were no approved closed events in this day on ' + site)
        except IndexError:
            print('IndexError: There were no approved closed events in this day on ' + site)

    return df_incidents_analysis

def analysis_active_incidents(site, index_site, df_incidents_analysis,df_active_events, df_info_sunlight):

    max_percentage = "{:.2%}".format(1)
    min_percentage = "{:.2%}".format(0)
    starttime_day = df_info_sunlight.at[index_site, 'Time of operation start']
    endtime_day = df_info_sunlight.at[index_site, 'Time of operation end']
    capacity_site = df_info_sunlight.at[index_site, 'Capacity']

    try:
        for index_event, row in df_active_events.iterrows():

            '''Reads each event'''
            if "roducing" in row['Component Status'] and not row['Curtailment Event'] == "x":

                starttime_event = df_active_events.loc[index_event, 'Rounded Event Start Time']
                capacity_affected = df_active_events.loc[index_event, 'Capacity Related Component']
                rel_comp = df_active_events.loc[index_event,'Related Component']

                if starttime_day > starttime_event:
                    starttime_event = starttime_day

                index_mint = df_incidents_analysis[
                    df_incidents_analysis['Time'] == starttime_event].index.values  # gets starting time row index
                int_index_mint = index_mint[0]  # turns index from numpy.ndarray to integer

                index_maxt = df_incidents_analysis[
                    df_incidents_analysis['Time'] == endtime_day].index.values  # gets ending time row index
                int_index_maxt = index_maxt[0]

                for index_timestamp in range(int_index_mint, int_index_maxt):

                    '''For each event read, it's effect is added in the corresponding period'''

                    percentage = "{:.2%}".format(capacity_affected / capacity_site)
                    if pd.isnull(df_incidents_analysis.loc[index_timestamp, site]):
                        percentage_final = "{:.2%}".format(1-(capacity_affected / capacity_site))
                        df_incidents_analysis.loc[index_timestamp, site] = percentage_final   #test1 max_percentage -
                    else:
                        perc_ce = float(df_incidents_analysis.loc[index_timestamp, site].strip('%')) / 100
                        perc_ae = float(percentage.strip('%')) / 100
                        percentage = "{:.2%}".format(perc_ce - perc_ae)


                        if float(percentage[:-1]) < float(min_percentage[:-1]):
                            percentage_final = min_percentage     #test1 - max_percentage
                        else:
                            percentage_final = percentage
                        df_incidents_analysis.loc[index_timestamp, site] = percentage_final   #test1 max_percentage -
    except KeyError:
        print('There were no approved active events in this day on ' + site)
    except NameError:
        print('There were no approved active events in this day on ' + site)
    except IndexError:
        print('There were no approved active events in this day on ' + site)


    return df_incidents_analysis

def analysis_closed_tracker_incidents(df_tracker_analysis, df_tracker_closed, df_info_sunlight):

    max_percentage = "{:.2%}".format(1)
    min_percentage = "{:.2%}".format(0)

    for index, row in df_tracker_closed.iterrows():
        #Site related info
        site = df_tracker_closed.loc[index, 'Site Name']
        index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site].index.values
        index_site = int(index_site_array[0])
        capacity_site = df_info_sunlight.at[index_site, 'Capacity']
        starttime_site = df_info_sunlight.loc[index_site, 'Time of operation start']
        endtime_site = df_info_sunlight.loc[index_site, 'Time of operation end']

        # Event related info
        starttime_event = df_tracker_closed.loc[index, 'Rounded Event Start Time']
        endtime_event = df_tracker_closed.loc[index, 'Rounded Event End Time']
        capacity_affected = df_tracker_closed.loc[index, 'Capacity Related Component']

        #print(type(starttime_event))
        #print(type(starttime_site))

        if starttime_event < starttime_site:
            print('Start time site: ' + str(starttime_site) + ' is later than Start time event' + str(starttime_event))
            starttime_event = starttime_site
        else:
            print(
                'Start time site: ' + str(starttime_site) + ' is earlier than Start time event' + str(starttime_event))

        if endtime_event > endtime_site:
            print('End time event: ' + str(endtime_event) + ' is later than End time site: ' + str(endtime_site))
            endtime_event = endtime_site
        else:
            print('End time event: ' + str(endtime_event) + ' is earlier than End time site: ' + str(endtime_site))

        index_mint = df_tracker_analysis[
            df_tracker_analysis['Time'] == starttime_event].index.values  # gets starting time row index
        int_index_mint = index_mint[0]  # turns index from numpy.ndarray to integer

        index_maxt = df_tracker_analysis[
            df_tracker_analysis['Time'] == endtime_event].index.values  # gets ending time row index
        int_index_maxt = index_maxt[0]

        for index in range(int_index_mint, int_index_maxt):
            percentage = "{:.2%}".format(capacity_affected / capacity_site)
            if pd.isnull(df_tracker_analysis.loc[index, site]):
                percentage_final = "{:.2%}".format(1-(capacity_affected / capacity_site))
                df_tracker_analysis.loc[index, site] = percentage_final   #test1 max_percentage -
            else:
                perc_ce = float(df_tracker_analysis.loc[index, site].strip('%')) / 100
                perc_ae = float(percentage.strip('%')) / 100
                percentage = "{:.2%}".format(perc_ce - perc_ae)
                if float(percentage[:-1]) < float(min_percentage[:-1]):
                    percentage_final = min_percentage     #test1 - max_percentage
                else:
                    percentage_final = percentage
                df_tracker_analysis.loc[index, site] = percentage_final   #test1 max_percentage -

    return df_tracker_analysis

def analysis_active_tracker_incidents(df_tracker_analysis, df_tracker_active, df_info_sunlight):

    max_percentage = "{:.2%}".format(1)
    min_percentage = "{:.2%}".format(0)

    for index, row in df_tracker_active.iterrows():
        # Site related info
        site = df_tracker_active.loc[index, 'Site Name']

        index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site].index.values
        index_site = int(index_site_array[0])

        capacity_site = df_info_sunlight.at[index_site, 'Capacity']
        starttime_site = df_info_sunlight.loc[index_site, 'Time of operation start']
        endtime_site = df_info_sunlight.loc[index_site, 'Time of operation end']

        # Event related info
        starttime_event = df_tracker_active.loc[index, 'Rounded Event Start Time']
        capacity_affected = df_tracker_active.loc[index, 'Capacity Related Component']

        if starttime_event < starttime_site:
            starttime_event = starttime_site
        try:
            index_mint = df_tracker_analysis[
                df_tracker_analysis['Time'] == starttime_event].index.values  # gets starting time row index
            int_index_mint = index_mint[0]  # turns index from numpy.ndarray to integer
        except IndexError:
            print("This event was not included because it went out of bounds in terms of start time")
            print(row)
        try:
            index_maxt = df_tracker_analysis[
                df_tracker_analysis['Time'] == endtime_site].index.values  # gets ending time row index
            int_index_maxt = index_maxt[0]
        except IndexError:
            print("This event was not included because it went out of bounds in terms of end time")
            print(row)

        for index in range(int_index_mint, int_index_maxt):
            percentage = "{:.2%}".format(capacity_affected / capacity_site)
            if pd.isnull(df_tracker_analysis.loc[index, site]):
                percentage_final = "{:.2%}".format(1-(capacity_affected / capacity_site))
                df_tracker_analysis.loc[index, site] = percentage_final   #test1 max_percentage -
            else:
                perc_ce = float(df_tracker_analysis.loc[index, site].strip('%')) / 100
                perc_ae = float(percentage.strip('%')) / 100
                percentage = "{:.2%}".format(perc_ce - perc_ae)
                if float(percentage[:-1]) < float(min_percentage[:-1]):
                    percentage_final = min_percentage     #test1 - max_percentage
                else:
                    percentage_final = percentage
                df_tracker_analysis.loc[index, site] = percentage_final   #test1 max_percentage -

    return df_tracker_analysis

def comprehensive_description(df):
    a=1
    return a

def describe_incidents(df,df_info_sunlight ,active_events: bool = False ,tracker: bool = False):

    if tracker == False:
        site_list = df.keys()
        for site in site_list:
            df_events = df[site]
            index_site_array = df_info_sunlight[df_info_sunlight['Site'] == site].index.values
            index_site = int(index_site_array[0])
            sunrise_time = df_info_sunlight.loc[index_site, 'Time of operation start']
            if active_events == False:
                print('Describing closed incidents of ' + site)
                for index, row in df_events.iterrows():
                    rel_comp = df_events.at[index, 'Related Component']
                    duration = df_events.at[index, 'Duration (h)']
                    start_date = df_events.at[index, 'Rounded Event Start Time']
                    end_date = df_events.at[index, 'Rounded Event End Time']
                    event_time_hour = end_date.hour
                    event_time_minute = end_date.minute
                    event_time = str(event_time_hour) + ':' + str(event_time_minute)
                    if start_date == sunrise_time and duration < 2:
                        description = "• " + str(rel_comp) + ' started late at ~' + str(event_time) + ' (closed)'
                    elif duration > 24:
                        description = "• " + str(rel_comp) + ' was not producing until ~' + str(event_time) + ' (closed)'
                    else:
                        description = "• " + str(rel_comp) + ' was not producing for ~' + str(duration) + ' hours (closed)'
                    df_events.loc[index, 'Comments'] = description

                df[site] = df_events

            else:
                print('Describing active incidents of ' + site)
                for index, row in df_events.iterrows():
                    rel_comp = df_events.at[index, 'Related Component']
                    start_date = df_events.at[index, 'Rounded Event Start Time']
                    day = start_date.day
                    month = start_date.month
                    date = str(day) + '/' + str(month)
                    description = "• " + str(rel_comp) + ' is not producing (open since ' + date  + ')'
                    df_events.loc[index, 'Comments'] = description

                df[site] = df_events
    else:

        if active_events == False:
            print('Describing closed tracker incidents')
            for index, row in df.iterrows():
                rel_comp = df.at[index, 'Related Component']
                duration = df.at[index, 'Duration (h)']
                description = "• " + str(rel_comp) + ' was off position for ~' + str(duration) + ' hours (closed)'
                df.loc[index, 'Comments'] = description
        else:
            print('Describing active tracker incidents')
            for index, row in df.iterrows():
                rel_comp = df.at[index, 'Related Component']
                start_date = df.at[index, 'Rounded Event Start Time']
                day = start_date.day
                month = start_date.month
                date = str(day) + '/' + str(month)
                description = "• " + str(rel_comp) + ' is off position (open since ' + date + ')'
                df.loc[index, 'Comments'] = description



    return df

def get_significance_score(df, active: bool = False):
    df_final = df
    if active == False:
        diff = (df['Event End Time'] - df['Event Start Time'])
        diff_days = [difference.days * (60 * 60 * 24) for difference in diff]
        diff_seconds = [difference.seconds for difference in diff]
        diff_total = [(diff_days[i] + diff_seconds[i]) / (60 * 60 * 24) for i in range(len(diff_seconds))]

        significance_score = [((df['Capacity Related Component'][i] * diff_total[i])/1000)
                              for i in range(len(diff_total))]


        df_final['Significance Score (MW*d)'] = significance_score

    else:
        today = datetime.today()
        diff = [today - df['Event Start Time'][i] for i in range(len(df['Event Start Time']))]
        diff_days = [difference.days * (60 * 60 * 24) for difference in diff]
        diff_seconds = [difference.seconds for difference in diff]
        diff_total = [(diff_days[i] + diff_seconds[i]) / (60 * 60 * 24) for i in range(len(diff_seconds))]

        significance_score = [((df['Capacity Related Component'][i] * diff_total[i]) / 1000) for i in
                              range(len(diff_total))]

        df_final['Significance Score (MW*d)'] = significance_score

    return df_final






















#Discontinued functions -------------------------------------------------------------------------------------------
def reset_final_report(Report_template_path, date, geography):
    year = date[:4]
    month = date[5:7]
    day = date[-2:]
    reportxl = openpyxl.load_workbook(Report_template_path)
    dir = os.path.dirname(Report_template_path)
    dir = dir.replace("/Info&Templates","")
    basename = os.path.basename(Report_template_path)

    reportfile = dir +  '/Reporting_'+ geography +'_Sites_' + str(day) + '-' + str(month) + '.xlsx'


    if not 'Active Events' in reportxl.sheetnames:
        pass
        print('Active Events sheet does not exist')
        reportxl.save(reportfile)
    elif reportxl['Active Events'].max_row == 1:
        pass
        print('Active Events sheet is empty')
        reportxl.save(reportfile)
    else:
        del reportxl['Active Events']
        reportxl.save(reportfile)
    # -----------------------------------------------------------
    if not 'Closed Events' in reportxl.sheetnames:
        pass
        print('Closed Events sheet does not exist')
        reportxl.save(reportfile)
    elif reportxl['Closed Events'].max_row == 1:
        pass
        print('Closed Events sheet is empty')
        reportxl.save(reportfile)
    else:
        del reportxl['Closed Events']
        reportxl.save(reportfile)
    # -----------------------------------------------------------
    if not 'Active tracker incidents' in reportxl.sheetnames:
        pass
        print('Active tracker events sheet does not exist')
        reportxl.save(reportfile)
    elif reportxl['Active tracker incidents'].max_row == 1:
        pass
        print('Active tracker events sheet is empty')
        reportxl.save(reportfile)
    else:
        del reportxl['Active tracker incidents']
        reportxl.save(reportfile)
    # -----------------------------------------------------------
    if not 'Closed tracker incidents' in reportxl.sheetnames:
        pass
        print('Closed tracker events sheet does not exist')
        reportxl.save(reportfile)
    elif reportxl['Closed tracker incidents'].max_row == 1:
        pass
        print('Closed tracker events sheet is empty')
        reportxl.save(reportfile)
    else:
        del reportxl['Closed tracker incidents']
        reportxl.save(reportfile)



    return reportfile

def impact_inv_blocks(df, df_impact_blocks_inv):
    for index, row in df.iterrows():
        component = row['Related Component']
        try:
            block = df_impact_blocks_inv.at[component, 'Block']
        except KeyError:
            block = ""
        if block:
            row['Comments'] = block
            df.loc[index, 'Comments'] = block
        else:
            print(component)

    for index, row in df.iterrows():
        component = row['Related Component']
        try:
            block = df_impact_blocks_inv.at[component, 'Block']
        except KeyError:
            block = ""
        if block:
            row['Comments'] = block
            df.loc[index, 'Comments'] = block
        else:
            print(component)
    return df

def remove_duplicates_dflist(df):
    ''' not done, add escape in case comp_type is not string '''
    comp_type = str(comp_type)
    remove_index = []
    component_list = []
    for index, row in df.iterrows():
        rel_comp = df.loc[index,'Related Component']
        if comp_type in rel_comp:
            remove_index.append(index)
            component_list.append(rel_comp)

    df_final = df.drop(remove_index)
    df_final = df_final.reset_index(None,drop = True)

    return df

def old_add_events_to_final_report(reportfile, df_list_active, df_list_closed,df_tracker_active, df_tracker_closed):
    x = 0
    for site in df_list_active.keys():
        df = df_list_active[site]
        if not df.empty:
            if x < 1:
                append_df_to_excel(reportfile, df_list_active[site], sheet_name='Active Events', startrow=0)
                print(site + ' active events added')
                x+=1
            else:
                append_df_to_excel_existing(reportfile,df_list_active[site], sheet_name= 'Active Events')
                print(site + ' active events added')
    x = 0
    for site in df_list_closed.keys():
        df = df_list_closed[site]
        if not df.empty:
            if x < 1:
                append_df_to_excel(reportfile, df_list_closed[site], sheet_name='Closed Events', startrow = 0)
                print(site + ' closed events added')
                x+=1
            else:
                append_df_to_excel_existing(reportfile, df_list_closed[site], sheet_name='Closed Events')
                print(site + ' closed events added')

    if not df_tracker_active.empty:
        append_df_to_excel(reportfile, df_tracker_active, sheet_name='Active tracker incidents', startrow=0)
        print('Tracker active events added')
    else:
        print('No tracker active events to be added')

    if not df_tracker_closed.empty:
        append_df_to_excel(reportfile, df_tracker_closed, sheet_name='Closed tracker incidents', startrow=0)
        print('Tracker closed events added')
    else:
        print('No tracker closed events to be added')

    return