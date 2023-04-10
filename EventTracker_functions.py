import PySimpleGUI as sg
import myprocesses as mp
import mysubprocesses as msp
import myfunctions as mf
import re
import os
import pandas as pd
import IPython
import datetime as dt
from datetime import datetime
import openpyxl
import xlsxwriter
import timeit
import math
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from scipy.optimize import curve_fit
import calendar
import seaborn as sns


def update_event_tracker(date_start,date_end,event_tracker_path,dmr_folder):

    if date_end == None:
        a=0
    else:
        a=1




    return a

def new_event_tracker_from_input():
    username = os.getlogin()

    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.


    layout = [[sg.Text('Choose the source of information:', pad=((2, 10), (2, 5)))],
              [sg.Radio('One file',group_id = "source", default = True, key = "-SRCOF-"), sg.Radio('DMR', group_id = "source", disabled = True, default = False, key = "-SRCDMR-")],
              [sg.Text('Select source of Desktop', pad=((0, 10), (10, 2)))],
              [sg.FolderBrowse(target='-SRCFOLDER-', initial_folder = "C:/Users/" + username + "/OneDrive - Lightsource BP/Desktop"),
               sg.In(key='-SRCFOLDER-', text_color='black', size=(20, 1), enable_events=True, readonly=True,
                     visible=True)],
              [sg.Text('Select report', pad=((0, 10), (10, 2)))],
              [sg.FileBrowse(target='-SRCFILE-',
                               initial_folder="C:/Users/" + username + "/OneDrive - Lightsource BP/Desktop"),
               sg.In(key='-SRCFILE-', text_color='black', size=(20, 1), enable_events=True, readonly=True,
                     visible=True)],
              [sg.Text('Enter geography ', pad=((0, 10), (10, 2))),sg.Push()],
              [sg.Combo(['AUS', 'ES', 'USA'], size=(4, 3), readonly=True, key='-GEO-', pad=((5, 10), (2, 10))),sg.Push(), sg.Checkbox('Recalculate All', enable_events=True, size=(13,3), pad=((20,0),(0,10)), key='chk_recalc')],
              [sg.Button('Submit'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Event Tracker', layout)

    toggle_sec1 = False
    toggle_updt = True
    toggle_recalc = False

    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read(timeout=100)

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            window.close()
            return "None","None", "None", "None", "None"
            break

        if event == 'chk_recalc':
            toggle_recalc = not toggle_recalc

        if event == 'Submit':

            source_folder = values['-SRCFOLDER-']
            source = values['-SRCFILE-']
            geography = values['-GEO-']
            geopgraphy_folder = source_folder + "/" + geography

            for key in values.keys():
                if "SRC" in key and values[key] == True:
                    if "OF" in key:
                        source_type = "one-file"
                    elif "DMR" in key:
                        source_type = "dmr"


            return source_folder,source, geography,geopgraphy_folder, toggle_recalc

            #print("Start date: ", date_start,"\n End date: ", date_end, "\n ET: ", event_tracker_path, "\n DMR folder: ", dmr_folder)

            """report_name = os.path.basename(Alarm_report_path)
            geography_report_match = re.search(r'\w+?_', report_name)
            geography_report = geography_report_match.group()[:-1]
            geography = values['-GEO-']"""

            #sg.popup_cancel('Under development')

    window.close()



    return

def collapse(layout, key, visible):
    """
    Helper function that creates a Column that can be later made hidden, thus appearing "collapsed"
    :param layout: The layout for the section
    :param key: Key used to make this section visible / invisible
    :param visible: visible determines if section is rendered visible or invisible on initialization
    :return: A pinned column that can be placed directly into your layout
    :rtype: sg.pin
    """
    return sg.pin(sg.Column(layout, key=key, visible=visible, pad=(0,0)))

def update_event_tracker_input():
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.

    end_date_calendar_section = [[sg.Text('Enter end date of period you want to add', pad=((2, 10), (2, 5)))],
              [sg.CalendarButton('Choose date', target='-ECAL-', format="%Y-%m-%d"),
               sg.In(key='-ECAL-', text_color='black', size=(16, 1), enable_events=True, readonly=True, visible=True)]]

    layout = [[sg.Text('Enter date of report you want to add', pad=((2, 10), (2, 5)))],
              [sg.CalendarButton('Choose date', target='-SCAL-', format="%Y-%m-%d"),
               sg.In(key='-SCAL-', text_color='black', size=(16, 1), enable_events=True, readonly=True, visible=True),
               sg.Checkbox('Multiple reports', enable_events=True, size=(13,1), key='chk_multr')],
              [collapse(end_date_calendar_section, '-EXCAL-', False)],
              [sg.Text('Choose Event Tracker to update', pad=((0, 10), (10, 2)))],
              [sg.FileBrowse(target='-ETFILE-'),
               sg.In(key='-ETFILE-', text_color='black', size=(20, 1), enable_events=True, readonly=True, visible=True)],
              [sg.Text('Choose location folder of DMRs', pad=((0, 10), (10, 2)))],
              [sg.FolderBrowse(target='-DMRFOLDER-'),
               sg.In(key='-DMRFOLDER-', text_color='black', size=(20, 1), enable_events=True, readonly=True,
                     visible=True)],
              [sg.Text('Enter geography ', pad=((0, 10), (10, 2))),sg.Push(), sg.Checkbox('Update All Export\n& Irradiance', default = True, enable_events=True, size=(13,3), pad=((20,0),(0,10)), key='chk_updt')],
              [sg.Combo(['AUS', 'ES', 'USA'], size=(4, 3), readonly=True, key='-GEO-', pad=((5, 10), (2, 10))),sg.Push(), sg.Checkbox('Recalculate All', enable_events=True, size=(13,3), pad=((20,0),(0,10)), key='chk_recalc')],
              [sg.Button('Submit'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Update Event Tracker', layout)

    toggle_sec1 = False
    toggle_updt = True
    toggle_recalc = False

    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read(timeout=100)

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            window.close()
            return "None", "None", "None", "None", "None","None", "None"
            break
        if event == 'chk_multr':
            toggle_sec1 = not toggle_sec1
            window['-EXCAL-'].update(visible=toggle_sec1)

        if event == 'chk_updt':
            toggle_updt = not toggle_updt

        if event == 'chk_recalc':
            toggle_recalc = not toggle_recalc

        if event == 'Submit':
            date_start = values['-SCAL-']  # date is string
            date_end = values['-ECAL-']
            event_tracker_path = values['-ETFILE-']
            dmr_folder = values['-DMRFOLDER-']
            geography = values['-GEO-']

            if date_end == "":
                date_end = None
                print('date end value changed to none')
            if toggle_sec1 == False:
                date_end = None

            return date_start, date_end, event_tracker_path, dmr_folder, geography,toggle_updt,toggle_recalc

            #print("Start date: ", date_start,"\n End date: ", date_end, "\n ET: ", event_tracker_path, "\n DMR folder: ", dmr_folder)

            """report_name = os.path.basename(Alarm_report_path)
            geography_report_match = re.search(r'\w+?_', report_name)
            geography_report = geography_report_match.group()[:-1]
            geography = values['-GEO-']"""

            #sg.popup_cancel('Under development')

    window.close()



    return

def event_tracker_from_input():
    username = os.getlogin()

    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.


    layout = [[sg.Text('Choose the source of information:', pad=((2, 10), (2, 5)))],
              [sg.Radio('Database',group_id = "source",disabled = True, default = False, key = "-SRCDB-"), sg.Radio('Event Tracker file', group_id = "source", default = True, key = "-SRCFILE-")],
              [sg.Text('Select source on Desktop', pad=((0, 10), (10, 2)))],
              [sg.FolderBrowse(target='-SRCFOLDER-', initial_folder = "C:/Users/" + username + "/OneDrive - Lightsource BP/Desktop"),
               sg.In(key='-SRCFOLDER-', text_color='black', size=(20, 1), enable_events=True, readonly=True,
                     visible=True)],
              [sg.Text('Enter geography ', pad=((0, 10), (10, 2))),sg.Push()],
              [sg.Combo(['AUS', 'ES', 'USA'], size=(4, 3), readonly=True, key='-GEO-', pad=((5, 10), (2, 10))),sg.Push(), sg.Checkbox('Recalculate All', enable_events=True, size=(13,3), pad=((20,0),(0,10)), key='chk_recalc')],
              [sg.Button('Submit'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Event Tracker', layout)

    toggle_sec1 = False
    toggle_updt = True
    toggle_recalc = False

    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read(timeout=100)

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            window.close()
            return "None", "None", "None", "None"
            break

        if event == 'chk_recalc':
            toggle_recalc = not toggle_recalc

        if event == 'Submit':

            source_folder = values['-SRCFOLDER-']
            geography = values['-GEO-']
            geopgraphy_folder = source_folder + "/" + geography

            for key in values.keys():
                if "SRC" in key and values[key] == True:
                    if "FILE" in key:
                        source_type = "file"
                    elif "DB" in key:
                        source_type = "database"


            return source_folder, geography,geopgraphy_folder, toggle_recalc

            #print("Start date: ", date_start,"\n End date: ", date_end, "\n ET: ", event_tracker_path, "\n DMR folder: ", dmr_folder)

            """report_name = os.path.basename(Alarm_report_path)
            geography_report_match = re.search(r'\w+?_', report_name)
            geography_report = geography_report_match.group()[:-1]
            geography = values['-GEO-']"""

            #sg.popup_cancel('Under development')

    window.close()



    return

def underperformance_report_input():
    username = os.getlogin()

    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.

    month_calendar_section = [[sg.Text('Enter month of analysis', pad=((2, 10), (2, 5)))],
                               [sg.CalendarButton('Choose date', target='-ECAL-', format="%Y-%m-%d"),
                                sg.In(key='-SCAL-', text_color='black', size=(16, 1), enable_events=True,
                                      readonly=True, visible=True)]]

    custom_calendar_section = [[sg.Text('Enter start date of period you want to analyse', pad=((2, 10), (2, 5)))],
                                 [sg.CalendarButton('Choose date', target='-SCAL-', format="%Y-%m-%d"),
                                  sg.In(key='-SCAL-', text_color='black', size=(16, 1), enable_events=True,
                                        readonly=True, visible=True)],
                             [sg.Text('Enter end date of period you want to analyse', pad=((2, 10), (2, 5)))],
                                 [sg.CalendarButton('Choose date', target='-ECAL-', format="%Y-%m-%d"),
                                  sg.In(key='-ECAL-', text_color='black', size=(16, 1), enable_events=True,
                                        readonly=True, visible=True)]]

    layout = [[sg.Text('Choose the source of information:', pad=((2, 10), (2, 5)))],
              [sg.Radio('Month',group_id = "period", default = False, key = "-PERMON-"), sg.Radio('Choose', group_id = "period", default = True, key = "-PERCHO-")],
              [sg.Text('Choose the period of analysis:', pad=((2, 10), (2, 5)))],
              [sg.Radio('Database', group_id="source", disabled=True, default=False, key="-SRCDB-"),
               sg.Radio('Event Tracker file', group_id="source", default=True, key="-SRCFILE-")],
              [sg.Text('Select source on Desktop', pad=((0, 10), (10, 2)))],
              [sg.FolderBrowse(target='-SRCFOLDER-', initial_folder = "C:/Users/" + username + "/OneDrive - Lightsource BP/Desktop"),
               sg.In(key='-SRCFOLDER-', text_color='black', size=(20, 1), enable_events=True, readonly=True,
                     visible=True)],
              [sg.Text('Enter geography ', pad=((0, 10), (10, 2))),sg.Push()],
              [sg.Combo(['AUS', 'ES', 'USA'], size=(4, 3), readonly=True, key='-GEO-', pad=((5, 10), (2, 10))),sg.Push(),
               sg.Checkbox('Recalculate All', enable_events=True, size=(13,3), pad=((20,0),(0,10)), key='chk_recalc')],
              [sg.Text('Select level of analysis', pad=((0, 10), (10, 2))),sg.Push(), sg.Text('Select Irradiance Threshold', pad=((0, 10), (10, 2))), sg.Push()],
              [sg.Combo(['All', 'Inverter level', 'Inverter only'], default_value = "All", size=(11, 3), readonly=True,
                        key='-LVL-', pad=((5, 10), (2, 10))),
               sg.Combo([20,50,85,100], default_value=50, size=(11, 3), readonly=True,key='-THR-', pad=((50, 10), (2, 10)))],
              [sg.Button('Submit'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Event Tracker', layout)

    toggle_sec1 = False
    toggle_updt = True
    toggle_recalc = False


    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read(timeout=100)

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            break

        if event == 'chk_recalc':
            toggle_recalc = not toggle_recalc

        if event == 'Submit':

            source_folder = values['-SRCFOLDER-']
            level = values['-LVL-']
            irradiance_threshold = values['-THR-']
            geography = values['-GEO-']
            geopgraphy_folder = source_folder + "/" + geography

            for key in values.keys():
                if "SRC" in key and values[key] == True:
                    if "FILE" in key:
                        source_type = "file"
                    elif "DB" in key:
                        source_type = "database"

                elif "PER" in key and values[key] == True:
                    if "CHO" in key:
                        period_list = ["choose"]
                    elif "MON" in key:
                        period_list = ["monthly"]


            return source_folder, geography,geopgraphy_folder, toggle_recalc, period_list,level, irradiance_threshold

            #print("Start date: ", date_start,"\n End date: ", date_end, "\n ET: ", event_tracker_path, "\n DMR folder: ", dmr_folder)

            """report_name = os.path.basename(Alarm_report_path)
            geography_report_match = re.search(r'\w+?_', report_name)
            geography_report = geography_report_match.group()[:-1]
            geography = values['-GEO-']"""

            #sg.popup_cancel('Under development')

    window.close()



    return

def mondaycom_file_input():
    sg.theme('DarkAmber')  # Add a touch of color
    # All the stuff inside your window.

    end_date_calendar_section = [[sg.Text('Enter end date of period you want to add', pad=((2, 10), (2, 5)))],
                                 [sg.CalendarButton('Choose date', target='-ECAL-', format="%Y-%m-%d"),
                                  sg.In(key='-ECAL-', text_color='black', size=(16, 1), enable_events=True,
                                        readonly=True, visible=True)]]

    layout = [[sg.Text('Enter date of report you want to add', pad=((2, 10), (2, 5)))],
              [sg.CalendarButton('Choose date', target='-SCAL-', format="%Y-%m-%d"),
               sg.In(key='-SCAL-', text_color='black', size=(16, 1), enable_events=True, readonly=True, visible=True),
               sg.Checkbox('Multiple reports', enable_events=True, size=(13, 1), key='chk_multr')],
              [collapse(end_date_calendar_section, '-EXCAL-', False)],
              [sg.Text('Choose location folder of Event Tracker', pad=((0, 10), (10, 2)))],
              [sg.FolderBrowse(target='-ETFOLDER-'),
               sg.In(key='-ETFOLDER-', text_color='black', size=(20, 1), enable_events=True, readonly=True,
                     visible=True)],
              [sg.Text('Enter geography ', pad=((0, 10), (10, 2))), sg.Push()],
              [sg.Combo(['AUS', 'ES', 'USA'], size=(4, 3), readonly=True, key='-GEO-', pad=((5, 10), (2, 10))),
               sg.Push()],
              [sg.Button('Submit'), sg.Exit()]]

    # Create the Window
    window = sg.Window('Monday.com files', layout)

    toggle_sec1 = False

    # Event Loop to process "events" and get the "values" of the inputs
    while True:
        event, values = window.read(timeout=100)

        if event == sg.WIN_CLOSED or event == 'Exit':  # if user closes window or clicks exit
            window.close()
            return "None", "None", "None", "None"
            break
        if event == 'chk_multr':
            toggle_sec1 = not toggle_sec1
            window['-EXCAL-'].update(visible=toggle_sec1)

        if event == 'Submit':
            date_start = values['-SCAL-']  # date is string
            date_end = values['-ECAL-']
            event_tracker_folder = values['-ETFOLDER-']
            geography = values['-GEO-']

            if date_end == "":
                date_end = date_start
            if toggle_sec1 == False:
                date_end = date_start

            return date_start, date_end, event_tracker_folder, geography,

            window.close()

            # print("Start date: ", date_start,"\n End date: ", date_end, "\n ET: ", event_tracker_path, "\n DMR folder: ", dmr_folder)

            """report_name = os.path.basename(Alarm_report_path)
            geography_report_match = re.search(r'\w+?_', report_name)
            geography_report = geography_report_match.group()[:-1]
            geography = values['-GEO-']"""

            # sg.popup_cancel('Under development')

    window.close()



    return

def get_files_to_add(date_start, date_end, dmr_folder, geography, no_update: bool = False):
    if no_update == False:
        if date_end == None:
            date_list = pd.date_range(date_start, date_start, freq = 'd')
        else:
            date_list = pd.date_range(date_start, date_end, freq = 'd')

        report_files_dict = {}
        irradiance_dict = {}
        export_dict = {}

        irradiance_folder = dmr_folder + "/Irradiance " + geography
        export_folder = dmr_folder + "/Exported Energy " + geography

        for date in date_list:
            #Get date info for name
            month = str("0" + str(date.month)) if date.month < 10 else str(date.month)
            day = str("0" + str(date.day)) if date.day < 10 else str(date.day)
            year = str(date.year)

            #Get each of the files to be used in each day
            report_file = dmr_folder + '/Reporting_' + geography + '_Sites_' + str(day) + "-" + str(month) + '.xlsx'  # will be picked by user
            irradiance_file = irradiance_folder + '/Irradiance_' + geography + '_Curated&Average-' + year + month + str(day) + '.xlsx'  # will be picked by user
            export_file = export_folder + '/Energy_Exported_' + geography + '_' + year + month + str(day) + '.xlsx'  # will be picked by user

            report_files_dict[date] = report_file
            irradiance_dict[date] = irradiance_file
            export_dict[date] = export_file

        report_files = list(report_files_dict.values())
        irradiance_files = list(irradiance_dict.values())
        export_files = list(export_dict.values())

        all_irradiance_file = irradiance_folder + '/All_Irradiance_' + geography + '.xlsx'  # will be picked by use
        all_export_file = export_folder + '/All_Energy_Exported_' + geography + '.xlsx'  # will be picked by use
        general_info_path = dmr_folder + '/Info&Templates/General Info ' + geography + '.xlsx'  # will be picked by script


        return report_files, irradiance_files, export_files, all_irradiance_file, all_export_file, general_info_path

    else:
        irradiance_folder = dmr_folder + "/Irradiance " + geography
        export_folder = dmr_folder + "/Exported Energy " + geography

        all_irradiance_file = irradiance_folder + '/All_Irradiance_' + geography + '.xlsx'  # will be picked by use
        all_export_file = export_folder + '/All_Energy_Exported_' + geography + '.xlsx'  # will be picked by use
        general_info_path = dmr_folder + '/Info&Templates/General Info ' + geography + '.xlsx'  # will be picked by script


        return all_irradiance_file, all_export_file, general_info_path

def get_general_info_dataframes(general_info_path):
    # Read general info file
    all_component_data = pd.read_excel(general_info_path, sheet_name='Component Code', engine='openpyxl')
    budget_irradiance = pd.read_excel(general_info_path, sheet_name='Budget Irradiance', index_col=0, engine='openpyxl')
    budget_pr = pd.read_excel(general_info_path, sheet_name='Budget PR', index_col=0, engine='openpyxl')
    budget_export = pd.read_excel(general_info_path, sheet_name='Budget Export', index_col=0, engine='openpyxl')

    # Separate data
    component_data = all_component_data.loc[
        (all_component_data['Component Type'] != 'Tracker') & (all_component_data['Component Type'] != 'Tracker Group')]
    tracker_data = all_component_data.loc[
        (all_component_data['Component Type'] == 'Tracker') | (all_component_data['Component Type'] == 'Tracker Group')]
    fmeca_data = pd.read_excel(general_info_path, sheet_name='FMECA', engine='openpyxl')

    site_capacities = component_data.loc[component_data['Component Type'] == 'Site'][
        ['Component', 'Nominal Power DC']].set_index('Component')
    fleet_capacity = site_capacities['Nominal Power DC'].sum()


    return component_data,tracker_data,fmeca_data,site_capacities,fleet_capacity,budget_irradiance,budget_pr,budget_export







#Calculation scripts

def calculate_active_hours_and_energy_lost(final_df_to_add,corrected_incidents_dict,df_all_irradiance, df_all_export,budget_pr,
                                           irradiance_threshold: int = 20, timestamp: int = 15, recalculate_value: bool=False):
    granularity = timestamp / 60
    tic = timeit.default_timer()
    for key, df in final_df_to_add.items():

        if "Active" in key:
            if "tracker" in key:
                continue
            else:
                active_events = True
                df = mf.rounddatesactive_15m("All", df)
                df = msp.calculate_activehours_energylost_incidents(df, df_all_irradiance, df_all_export, budget_pr,
                                                                    corrected_incidents_dict, active_events,
                                                                    recalculate_value, granularity)

                df = mf.match_df_to_event_tracker(df, None, None, active=active_events, simple_match=True)
                final_df_to_add[key] = df


        elif "Closed" in key:
            if "tracker" in key:
                continue
            else:
                active_events = False
                df = msp.calculate_activehours_energylost_incidents(df, df_all_irradiance, df_all_export, budget_pr,
                                                                    corrected_incidents_dict, active_events,
                                                                    recalculate_value, granularity)
                df = mf.match_df_to_event_tracker(df, None, None, simple_match=True)
                final_df_to_add[key] = df

        else:
            continue

    toc = timeit.default_timer()
    print(toc - tic)

    return final_df_to_add

def calculate_availability_in_period(incidents, period, component_data, df_all_irradiance, df_all_export,budget_pr,
                                     irradiance_threshold: int = 20,timestamp: int = 15):
    granularity = timestamp / 60

    #Get dates from period info
    date_start_str, date_end_str = msp.choose_period_of_analysis(period)
    date_range = date_start_str + " to " + date_end_str
    print(date_range)


    #Get site list --------- could be input
    site_list = list(set([re.search(r'\[.+\]', site).group().replace('[', "").replace(']', "") for site in
                          df_all_irradiance.loc[:, df_all_irradiance.columns.str.contains('Irradiance')].columns]))
    site_list = [mf.correct_site_name(site) for site in site_list]


    #Get site and fleet capacities --------- could be input
    site_capacities = component_data.loc[component_data['Component Type'] == 'Site'][
        ['Component', 'Nominal Power DC']].set_index('Component').loc[site_list,:]
    fleet_capacity = site_capacities['Nominal Power DC'].sum()

    #Get only incidents that count for availaility, aka, "Not producing"
    incidents = incidents.loc[incidents['Component Status'] == "Not Producing"].reset_index(None, drop=True)

    # Calculate Availability, Active Hours and Corrected Dataframe
    availability_period_per_site = {}
    active_hours_per_site = {}
    incidents_corrected_period_per_site = {}

    for site in site_list:
        availability_period, site_active_hours_period, corrected_relevant_incidents = mp.calculate_availability_period(
            site, incidents, component_data, budget_pr, df_all_irradiance, df_all_export, irradiance_threshold,
            date_start_str, date_end_str, granularity)

        availability_period_per_site[site] = availability_period
        active_hours_per_site[site] = site_active_hours_period
        incidents_corrected_period_per_site[site] = corrected_relevant_incidents

    #Add fleet value and company goals values
    availability_period_per_site['Fleet'] = sum(
            [availability_period_per_site[site] * site_capacities.loc[site, 'Nominal Power DC'] for site in
             site_list]) / fleet_capacity
    availability_period_per_site['Company goal'] = 0.944
    availability_period_per_site['Company max goal'] = 0.964



    availability_period_df = pd.DataFrame.from_dict(availability_period_per_site, orient='index', columns=[
        date_range])  # , orient='index', columns=['Incident weighted downtime (h)'])
    activehours_period_df = pd.DataFrame.from_dict(active_hours_per_site, orient='index', columns=[date_range])
    incidents_corrected_period = pd.concat(list(incidents_corrected_period_per_site.values()))






    return availability_period_df,activehours_period_df,incidents_corrected_period,date_range


def calculate_pr_in_period(incidents_period,availability_period, period, component_data, df_all_irradiance, df_all_export,budget_pr,budget_export,
                                     budget_irradiance, irradiance_threshold: int = 20,timestamp: int = 15):

    # Get site list --------- could be input
    site_list = list(set([re.search(r'\[.+\]', site).group().replace('[', "").replace(']', "") for site in
                          df_all_irradiance.loc[:, df_all_irradiance.columns.str.contains('Irradiance')].columns]))
    site_list = [mf.correct_site_name(site) for site in site_list]

    # Get site and fleet capacities --------- could be input
    site_capacities = component_data.loc[component_data['Component Type'] == 'Site'][
                          ['Component', 'Nominal Power DC']].set_index('Component').loc[site_list, :]
    fleet_capacity = site_capacities['Nominal Power DC'].sum()


    #Get dates from period info
    date_start_str, date_end_str = msp.choose_period_of_analysis(period)

    # Get start time analysis
    date_start_avail_analysis = datetime.strptime(date_start_str, '%Y-%m-%d').date()
    timestamp_start_avail_analysis = datetime.strptime(date_start_str + " 00:00:00", '%Y-%m-%d %H:%M:%S')

    # Get end time analysis
    date_end_avail_analysis = datetime.strptime(date_end_str, '%Y-%m-%d').date()
    date_end_str_event = str(datetime.strptime(date_end_str, '%Y-%m-%d').date() + dt.timedelta(days=1))
    timestamp_end_avail_analysis = datetime.strptime(date_end_str_event + " 00:00:00", '%Y-%m-%d %H:%M:%S')

    # Get Data to analyse: incidents, export data and irradiance data
    df_export_period = df_all_export.loc[(df_all_export['Timestamp'] >= timestamp_start_avail_analysis) & (
                df_all_export['Timestamp'] <= timestamp_end_avail_analysis)].set_index('Timestamp')

    df_irradiance_period = df_all_irradiance.loc[(df_all_irradiance['Timestamp'] >= timestamp_start_avail_analysis) & (
                df_all_irradiance['Timestamp'] <= timestamp_end_avail_analysis)].set_index('Timestamp')

    print(date_start_avail_analysis, " ", date_end_avail_analysis)

    pr_period_per_site = {}
    data_period_per_site = {}

    for site in site_list:
        print(site)
        export_column = list(df_export_period.columns[df_export_period.columns.str.contains(str(site))].values)[0]
        site_capacity = site_capacities.loc[site, "Nominal Power DC"]

        if len(export_column) > 0:
            export_data = df_export_period[[export_column]]
            exported_energy = float(export_data.sum())

            # Get relevant irradiance data and calculate expected energy
            irradiance_site = df_irradiance_period.loc[:, df_irradiance_period.columns.str.contains(site)]
            actual_column, curated_column, data_gaps_proportion, poa_avg_column = mf.get_actual_irradiance_column(
                irradiance_site)

            if not actual_column == None:
                actual_irradiance_site = irradiance_site.loc[:, actual_column]
            else:
                actual_irradiance_site = irradiance_site.loc[:, poa_avg_column]

            start_timestamp = actual_irradiance_site.index[0]
            end_timestamp = actual_irradiance_site.index[-1]
            start_day = start_timestamp.date()
            end_day = actual_irradiance_site.index[-2].date()

            # Calculate Expected Energy in period
            expected_energy, expected_energy_info = mf.calculate_expected_energy(site, start_timestamp, end_timestamp,
                                                                                 budget_export, budget_irradiance,
                                                                                 actual_irradiance_site)
            if "Bighorn" in site:
                """print(start_timestamp, "\n", end_timestamp, "\n",expected_energy, "\n" )"""
                for key in expected_energy_info.keys():
                    print(key, ": ", expected_energy_info[key]["Expected Energy Period"], "\n")

            # Calculate Energy Lost
            energy_lost = incidents_period.loc[incidents_period['Site Name'] == site][
                              'Energy Lost (MWh)'].replace("", 0).sum() * 1000
            """try:
                energy_lost = incidents_period.loc[incidents_period['Site Name'] == site][
                                  'Energy Lost (MWh)'].replace("", 0).sum() * 1000

                print(incidents_period.loc[incidents_period['Site Name'] == site])
                print(energy_lost)
                print(type(energy_lost))
            except TypeError:
                print("TypeError, replacing nan values")
                energy_lost = incidents_period.loc[incidents_period['Site Name'] == site]['Energy Lost (MWh)'].replace(
                    "", 0).sum() * 1000"""

            # Calculate PRs
            actual_pr = exported_energy / ((actual_irradiance_site.sum() / 4000) * site_capacity)
            possible_pr = (exported_energy + energy_lost) / ((actual_irradiance_site.sum() / 4000) * site_capacity)

            # Calculate Variances
            actual_expected_variance = (exported_energy / expected_energy) - 1
            corrected_actual_expected_variance = ((exported_energy + energy_lost) / expected_energy) - 1

            # Get availability in Period
            availability = availability_period.loc[site, :].values[0]

            # print(site, "\n Expected Energy: ", expected_energy,"\n Exported Energy: " ,exported_energy, "\n Energy Lost: ", energy_lost)
            # incidents_period_site = incidents_period.loc[(incidents_period['Component Status'] == "Not Producing") & (incidents['Site Name'] == site) & (incidents['Event Start Time'] > )].reset_index(None, drop=True)

            # Store relevant data

            data_period_per_site[site] = ("{:.2%}".format(availability),
                                          "{:.2%}".format(actual_pr),
                                          "{:,.2f}".format(exported_energy),
                                          "{:,.2f}".format(energy_lost),
                                          "{:.2%}".format(possible_pr),
                                          "{:,.2f}".format(actual_irradiance_site.sum() / 4000),
                                          "{:,.2f}".format(expected_energy),
                                          "{:.2%}".format(actual_expected_variance),
                                          "{:.2%}".format(corrected_actual_expected_variance),
                                          "{:.2%}".format(data_gaps_proportion))


        else:
            print(site, " Exported energy data not found.")
            continue

    data_period_df = pd.DataFrame.from_dict(data_period_per_site, columns=['Availability (%)', 'Actual PR (%)',
                                                                           "Actual Exported Energy (kWh)",
                                                                           "Energy Lost (kWh)", "Corrected PR (%)",
                                                                           "Actual Irradiance (kWh/m2)",
                                                                           "Expected Energy (kWh)",
                                                                           "Actual vs Expected Energy Variance",
                                                                           "Corrected Actual vs Expected Energy Variance",
                                                                           "Data Gaps (%)"], orient='index')

    return data_period_df



def availability_visuals(availability_fleet_per_period, period, folder_img):

    df = availability_fleet_per_period[period]
    df.index = df.index.astype('str')
    df_to_plot = df[df.index.str.contains('LSBP')].sort_index()
    df_to_plot_line = df[~df.index.str.contains('LSBP')]  # variable is a Series not a df

    if period == 'choose':
        title = df_to_plot.columns.to_list()[0]
    else:
        title = period

    # month = 9

    x_labels = df_to_plot.columns.to_list()
    y_values_labels = df_to_plot.index.to_list()
    # y_values_labels = [name.replace('Corrected (w/clipping) Monthly', '').replace(' PR %','') for name in y_values_labels]
    colors = ['#FE5000' if "LSBP" in name else '#FF5353' for name in y_values_labels]
    y_values = df_to_plot[x_labels[0]] * 100
    y_values_lines = df_to_plot_line[x_labels[0]] * 100

    plt.figure(figsize=(27, 9))
    plt.style.use('ggplot')
    plt.suptitle(str(title.upper()) + ' Availability %', fontsize='xx-large')
    plt.ylabel('Availability %', fontsize='xx-large')
    plt.bar(y_values_labels, y_values, width=0.6, color=colors)
    plt.xticks(rotation=45, ha='right', fontsize='xx-large')
    plt.yticks(fontsize='xx-large')
    plt.ylim([0, 100])
    for index, data in enumerate(y_values):
        label = str(data)[:5] + "%"
        plt.text(x=index - 0.25, y=data + 1, s=label, fontdict=dict(fontsize=18))

    # Fleet line
    plt.axhline(y=y_values_lines['Fleet'], linewidth=2, color='black', linestyle='-.', label='Fleet')
    plt.text(-1.4, y_values_lines['Fleet'], s="{:.2%}".format(y_values_lines['Fleet'] / 100),
             fontdict=dict(fontsize=15))

    # Company goal line
    plt.axhline(y=y_values_lines['Company goal'], linewidth=2, color='red', linestyle='-.', label='Company goal')
    plt.text(-1.4, y_values_lines['Company goal'], s="{:.2%}".format(y_values_lines['Company goal'] / 100),
             fontdict=dict(fontsize=15))

    # Company max goal line
    plt.axhline(y=y_values_lines['Company max goal'], linewidth=2, color='green', linestyle='-.',
                label='Company max goal')
    plt.text(-1.4, y_values_lines['Company max goal'], s="{:.2%}".format(y_values_lines['Company max goal'] / 100),
             fontdict=dict(fontsize=15))

    plt.legend()

    period_graph = (folder_img + '/' + str(period.upper()) + '_availability.png')
    plt.savefig(period_graph, bbox_inches='tight')


    return period_graph


def create_event_tracker_file_all(final_df_to_add, dest_file,performance_fleet_per_period, site_capacities,
                                  dict_fmeca_shapes):

    writer = pd.ExcelWriter(dest_file, engine='xlsxwriter', engine_kwargs={'options': {'nan_inf_to_errors': True}})
    workbook = writer.book


    # <editor-fold desc="Formats">
    # Format column header
    format_darkblue_white = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'bg_color': '#002060', 'font_color': '#FFFFFF'})
    format_darkblue_white.set_bold()
    format_darkblue_white.set_text_wrap()

    format_lightblue_black = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'font_color': '#000000'})
    format_lightblue_black.set_bold()
    format_lightblue_black.set_text_wrap()
    format_lightblue_black.set_border()

    format_header = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'font_color': '#000000'})
    format_header.set_bold()
    format_header.set_text_wrap()

    format_all_white = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})
    format_all_black = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#000000'})
    format_black_on_white = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#FFFFFF'})

    # Format of specific column data
    format_day_data = workbook.add_format({'num_format': 'dd/mm/yyyy', 'valign': 'vcenter'})
    format_day_data.set_align('right')
    format_day_data.set_border()

    format_hour_data = workbook.add_format({'num_format': 'hh:mm:ss', 'valign': 'vcenter'})
    format_hour_data.set_align('right')
    format_hour_data.set_border()

    format_day_hour = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm:ss', 'valign': 'vcenter'})
    format_day_hour.set_align('right')
    format_day_hour.set_border()

    # Format numbers
    format_number = workbook.add_format({'num_format': '#,##0.00', 'align': 'center', 'valign': 'vcenter'})
    format_number.set_border()

    format_nodecimal = workbook.add_format({'num_format': '0', 'align': 'center', 'valign': 'vcenter'})
    format_nodecimal.set_border()

    format_percentage = workbook.add_format({'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter'})
    format_percentage.set_border()

    format_percentage_good = workbook.add_format(
            {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#C6EFCE',
             'font_color': '#006100'})
    format_percentage_good.set_border()
    format_percentage_mid = workbook.add_format(
            {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFEB9C',
             'font_color': '#9C5700'})
    format_percentage_mid.set_border()
    format_percentage_bad = workbook.add_format(
            {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFC7CE',
             'font_color': '#9C0006'})
    format_percentage_bad.set_border()

    # Format strings
    format_string = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string.set_border()

    format_string_wrapped = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string_wrapped.set_text_wrap()
    format_string_wrapped.set_border()

    format_string_unlocked = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'locked': False})
    unlocked = workbook.add_format({'locked': False})
    format_string_unlocked.set_border()

    format_string_bold = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold.set_bold()
    format_string_bold.set_border()

    format_string_bold_wrapped = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold_wrapped.set_bold()
    format_string_bold_wrapped.set_border()
    format_string_bold_wrapped.set_text_wrap()

    format_first_column = workbook.add_format(
            {'align': 'center', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#000000'})
    format_first_column.set_bold()
    format_first_column.set_border()
    format_first_column.set_text_wrap()
    # </editor-fold>

    # <editor-fold desc="YTD Performance Overview Sheet">
    sheet = "YTD Performance Overview"
    try:
        ws_sheet = workbook.add_worksheet(sheet)
    except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
        sheet = sheet + "_new"
        ws_sheet = workbook.add_worksheet(sheet)

    df_performance = performance_fleet_per_period['ytd'].T

    sites = list(df_performance.columns)

    start_row_header = 1
    start_row_data = 2
    start_column = 0

    for site in sites:
        level = 0
        start_row_header_str = str(start_row_header)
        start_row_data_str = str(start_row_data)

        performance_site = df_performance.loc[:, [site]].reset_index()
        n_rows_performance = performance_site.shape[0] + 1
        n_columns_performance = performance_site.shape[1]

        df_total = performance_site

        max_rows = n_rows_performance
        n_columns_total = df_total.shape[1]

        width = mf.get_col_widths(df_total)

        print("\n", df_total)

        for i in range(start_column, start_column + n_columns_total):

            header = df_total.columns[i - start_column]

            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + start_row_header_str
            data_cell = column_letter + start_row_data_str
            all_column = column_letter + ':' + column_letter

            # print('Header: ', header, "\n", 'Header cell:', header_cell, "\n", "Data Cell: ", data_cell ,"\n")

            data = list(df_total[header].fillna(""))

            if header == "index":
                to_colapse_column1 = column_letter
                data = [x for x in data if not pd.isnull(x)]
                ws_sheet.write(header_cell, "", format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_lightblue_black)
                if column_letter == "A":
                    ws_sheet.set_column(all_column, 23)
                else:
                    ws_sheet.set_column(all_column, 0)  # ,None,{'level': 1, 'hidden': True})

            elif "LSBP" in header or "Wellington" in header:
                kpis = df_total['index']
                ws_sheet.write(header_cell, header, format_darkblue_white)
                data = [x for x in data if not x == ""]

                for i in range(len(data)):
                    cell = column_letter + str(start_row_data + i)
                    value = data[i]
                    kpi = kpis[i]

                    if "%" in value:
                        value = float(value[:-1]) / 100
                        ws_sheet.write_number(cell, value, format_percentage)
                        if not "PR (%)" in kpi:
                            if "Availability" in kpi:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [{'criteria': '>=', 'type': 'number',
                                                                              'value': 0.97},
                                                                             {'criteria': '<', 'type': 'number',
                                                                              'value': 0.10},
                                                                             {'criteria': '<=', 'type': 'number',
                                                                              'value': 0.10}]})

                            else:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [
                                                                       {'criteria': '<=', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>=', 'type': 'number',
                                                                        'value': 0.05}]})


                    else:
                        value = float(value.replace(",", ""))
                        ws_sheet.write_number(cell, value, format_number)

                ws_sheet.set_column(all_column, 16, None)


            else:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, 18, None)  # ,{'level': 1, 'hidden': True})

        level = level + 1
        start_column = start_column + n_columns_total
    # </editor-fold>

    # <editor-fold desc="MTD Performance Overview Sheet">
    active_events = final_df_to_add['Active Events']
    overview_events = active_events.loc[active_events['Component Status'] == "Not Producing"][
        ['Site Name', 'ID', 'Related Component', 'Event Start Time', 'Energy Lost (MWh)', 'Capacity Related Component']]
    overview_events['% of site affected'] = [
        "{:.2%}".format(row['Capacity Related Component'] / float(site_capacities.loc[row['Site Name']])) for index, row
        in overview_events.iterrows()]
    overview_events['Actions'] = active_events.loc[active_events['Component Status'] == "Not Producing"]['Remediation']
    overview_events['Space'] = ""
    overview_events

    sheet = "MTD Performance Overview"
    try:
        ws_sheet = workbook.add_worksheet(sheet)
    except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
        sheet = sheet + "_new"
        ws_sheet = workbook.add_worksheet(sheet)

    try:
        df_performance = performance_fleet_per_period['mtd'].T
    except KeyError:
        df_performance = performance_fleet_per_period['monthly'].T

    sites = list(df_performance.columns)

    start_row_header = 1
    start_row_data = 2
    start_column = 0

    for site in sites:
        level = 0
        start_row_header_str = str(start_row_header)
        start_row_data_str = str(start_row_data)

        performance_site = df_performance.loc[:, [site]].reset_index()
        n_rows_performance = performance_site.shape[0] + 1
        n_columns_performance = performance_site.shape[1]

        incidents_site = overview_events.loc[overview_events['Site Name'] == site].reset_index(drop=True)
        # incidents_site.insert(1, "#", list(range(1,incidents_site.shape[0] + 1)))
        n_rows_incidents = incidents_site.shape[0] + 1
        n_columns_incidents = incidents_site.shape[1]

        df_total = pd.concat([performance_site, incidents_site], axis=1)

        max_rows = max(n_rows_performance, n_rows_incidents)
        n_columns_total = df_total.shape[1]

        width = mf.get_col_widths(df_total)

        #print("\n", df_total)

        for i in range(start_column, start_column + n_columns_total):

            header = df_total.columns[i - start_column]

            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + start_row_header_str
            data_cell = column_letter + start_row_data_str
            all_column = column_letter + ':' + column_letter

            # print('Header: ', header, "\n", 'Header cell:', header_cell, "\n", "Data Cell: ", data_cell ,"\n")

            data = list(df_total[header].fillna(""))

            if header == "index":
                to_colapse_column1 = column_letter
                data = [x for x in data if not pd.isnull(x)]
                ws_sheet.write(header_cell, "", format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_lightblue_black)
                if column_letter == "A":
                    ws_sheet.set_column(all_column, 23)
                else:
                    ws_sheet.set_column(all_column, 23 ,None,{'level': 1, 'hidden': True})  # ,None,{'level': 1, 'hidden': True})

            elif "LSBP" in header or "Wellington" in header:
                kpis = df_total['index']
                ws_sheet.write(header_cell, header, format_darkblue_white)
                data = [x for x in data if not x == ""]

                for i in range(len(data)):
                    cell = column_letter + str(start_row_data + i)
                    value = data[i]
                    kpi = kpis[i]

                    if "%" in value:
                        value = float(value[:-1]) / 100
                        ws_sheet.write_number(cell, value, format_percentage)
                        if not "PR (%)" in kpi:
                            if "Availability" in kpi:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [{'criteria': '>=', 'type': 'number',
                                                                              'value': 0.97},
                                                                             {'criteria': '<', 'type': 'number',
                                                                              'value': 0.10},
                                                                             {'criteria': '<=', 'type': 'number',
                                                                              'value': 0.10}]})

                            else:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [
                                                                       {'criteria': '<=', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>=', 'type': 'number',
                                                                        'value': 0.05}]})


                    else:
                        value = float(value.replace(",", ""))
                        ws_sheet.write_number(cell, value, format_number)

                ws_sheet.set_column(all_column, 16, None)


            elif "Time" in header:
                data = [x for x in data if not pd.isnull(x)]
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_day_hour)
                ws_sheet.set_column(all_column, 20, None, {'level': 1, 'hidden': True})

            elif "%" in header:
                to_colapse_column2 = column_letter
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_percentage)
                ws_sheet.set_column(all_column, 15, None, {'level': 1, 'hidden': True})

            elif "Capacity" in header or "(" in header:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_number)
                ws_sheet.set_column(all_column, 15, None, {'level': 1, 'hidden': True})

            elif "ID" in header:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string_bold_wrapped)
                ws_sheet.set_column(all_column, 18, None, {'level': 1, 'hidden': True})



            elif "Site Name" in header:
                data = list(range(50))
                ws_sheet.write(header_cell, "", format_all_white)
                ws_sheet.write_column(data_cell, data, format_all_white)
                ws_sheet.set_column(all_column, 1, None, {'level': 1, 'hidden': True})

            elif "Space" in header:
                to_colapse_column = column_letter
                data = list(range(50))
                ws_sheet.write(header_cell, "+", format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_all_white)
                ws_sheet.set_column(all_column, 2, None, {'collapsed': True})

            elif "Actions" in header:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)  # format_string_wrapped
                ws_sheet.set_column(all_column, 55, None, {'level': 1, 'hidden': True})


            else:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, 18, None, {'level': 1, 'hidden': True})

        level = level + 1
        start_column = start_column + n_columns_total
    # </editor-fold>

    # <editor-fold desc="FMECA AUX sheet">
    start_row_index = 1
    start_column_index = 1
    start_column = openpyxl.utils.cell.get_column_letter(1)
    dict_fmeca_table_range = {}
    for name, data in dict_fmeca_shapes.items():
        df = data[0]
        shape = data[1]

        n_row = shape[0]
        n_column = shape[1]

        start_column = openpyxl.utils.cell.get_column_letter(1)
        end_column = openpyxl.utils.cell.get_column_letter(shape[1])
        end_row = start_row_index + n_row

        start_cell = start_column + str(start_row_index)
        table_range = "$" + start_column + "$" + str(start_row_index + 1) + ":$" + end_column + "$" + str(end_row)
        dict_fmeca_table_range[name] = table_range

        # range
        # print(df)

        df.to_excel(writer, sheet_name='FMECA_AUX', startrow=start_row_index - 1, startcol=start_column_index - 1,
                    index=False)

        for i in range(len(df.columns)):
            range_name = df.columns[i]
            # print(range_name)
            column = openpyxl.utils.cell.get_column_letter(i + 1)
            range_cells = '$' + column + "$" + str(start_row_index + 1) + ":$" + column + "$" + str(end_row)
            workbook.define_name(range_name, '=FMECA_AUX!' + range_cells)
            """if "ategory" not in name:
                workbook.define_name(range_name, '=FMECA_AUX!' + range_cells)"""

        # Prepare next iteration
        start_row_index = start_row_index + n_row + 2
    # </editor-fold>

    # <editor-fold desc="Events' sheets">
    fmeca_columns = final_df_to_add['FMECA'].columns.to_list()
    n_rows_fmeca = final_df_to_add['FMECA'].shape[0]
    n_columns_fmeca = final_df_to_add['FMECA'].shape[1]
    reference_column = openpyxl.utils.cell.get_column_letter(
        final_df_to_add['FMECA'].columns.to_list().index('Fault') + 1)

    for sheet in final_df_to_add.keys():
        df = final_df_to_add[sheet]
        width = mf.get_col_widths(df)
        n_rows = df.shape[0]
        n_columns = df.shape[1]
        try:
            ws_sheet = workbook.add_worksheet(sheet)
        except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
            sheet = sheet + "_new"
            ws_sheet = workbook.add_worksheet(sheet)
        if "Closed" in sheet or "Active" in sheet:
            for i in range(len(df.columns)):
                header = df.columns[i]
                column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
                header_cell = column_letter + '1'
                data_cell = column_letter + '2'
                all_column = column_letter + ':' + column_letter
                data = df[header].fillna("")

                if header == 'ID':
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_first_column)
                    ws_sheet.set_column(all_column, 18)

                elif "Time" in header:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_day_hour)
                    ws_sheet.set_column(all_column, 19)

                elif "Capacity" in header or "(" in header:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_number)
                    ws_sheet.set_column(all_column, 12)

                elif "Fa" in header or "ategory" in header:
                    if header == "Resolution Category":
                        ws_sheet.write(header_cell, header, format_header)
                        ws_sheet.write_column(data_cell, data, format_string_unlocked)
                        ws_sheet.set_column(all_column, width[i + 1], unlocked)
                        ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                 {'validate': 'list', 'source': ['Repair', 'Reset', 'Part Replacement', 'Unit Replacement']})
                    else:
                        fmeca_column_match = openpyxl.utils.cell.get_column_letter(fmeca_columns.index(header) + 1)
                        ws_sheet.write(header_cell, header, format_header)
                        ws_sheet.write_column(data_cell, data, format_string_unlocked)
                        ws_sheet.set_column(all_column, width[i + 1], unlocked)

                        # Add Data validation
                        if header == 'Fault':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=FMECA_AUX!' + str(dict_fmeca_table_range['Faults'])})
                            fault_cell = data_cell

                        elif header == 'Fault Component':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' + fault_cell + ', " ", "_"), "-","_"))'})
                            fcomp_cell = data_cell

                        elif header == 'Failure Mode':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' + fault_cell + '&"_"&' + fcomp_cell + '," ", "_"),"-","_"))'})
                            fmode_cell = data_cell

                        elif header == 'Failure Mechanism':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' + fault_cell + '&"_"&' + fcomp_cell + '&"_"&' + fmode_cell + ', " ", "_"), "-","_"))'})
                            fmec_cell = data_cell

                        elif header == 'Category':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' + fault_cell + '&"_"&' + fcomp_cell + '&"_"&' + fmode_cell + '&"_"&' + fmec_cell + ', " ", "_"), "-","_"))'})
                            cat_cell = data_cell
                        elif header == 'Subcategory':
                            ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                                     {'validate': 'list',
                                                      'source': '=INDIRECT(SUBSTITUTE(SUBSTITUTE(' + fault_cell + '&"_"&' + fcomp_cell + '&"_"&' + fmode_cell + '&"_"&' + fmec_cell + '&"_"&' + cat_cell + ', " ", "_"), "-","_"))'})
                            subcat_cell = data_cell



                elif header == "Incident Status":
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string_unlocked)
                    ws_sheet.set_column(all_column, width[i + 1], unlocked)
                    ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                             {'validate': 'list', 'source': ['Open', 'Closed']})

                elif header == "Categorization Status":
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string_unlocked)
                    ws_sheet.set_column(all_column, width[i + 1], unlocked)
                    ws_sheet.data_validation(data_cell + ":" + data_cell[0] + str(1 + n_rows),
                                             {'validate': 'list', 'source': ['Pending', 'Completed']})


                elif header == 'Remediation' or header == 'Comments':
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string_wrapped)
                    ws_sheet.set_column(all_column, 60)


                else:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string)
                    ws_sheet.set_column(all_column, width[i + 1])
        else:
            for i in range(len(df.columns)):
                header = df.columns[i]
                column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
                header_cell = column_letter + '1'
                data_cell = column_letter + '2'
                all_column = column_letter + ':' + column_letter
                data = df[header].fillna("")

                if "ID" in header:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_first_column)
                    ws_sheet.set_column(all_column, width[i + 1])
                else:
                    ws_sheet.write(header_cell, header, format_header)
                    ws_sheet.write_column(data_cell, data, format_string)
                    ws_sheet.set_column(all_column, width[i + 1])

        ws_sheet.set_default_row(30)
    # </editor-fold>

    ws_active = workbook.get_worksheet_by_name("MTD Performance Overview")
    ws_active.activate()

    ws_fmeca_aux = workbook.get_worksheet_by_name('FMECA_AUX')
    ws_fmeca_aux.hide()

    writer.save()
    print('Done')

    return

def create_underperformance_report(underperformance_dest_file,incidents_corrected_period, performance_fleet_per_period):

    writer_und = pd.ExcelWriter(underperformance_dest_file, engine='xlsxwriter',engine_kwargs={'options': {'nan_inf_to_errors': True}})
    workbook = writer_und.book

    # <editor-fold desc="Formats">
    # Format column header
    format_darkblue_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#002060', 'font_color': '#FFFFFF'})
    format_darkblue_white.set_bold()
    format_darkblue_white.set_text_wrap()

    format_lightblue_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#DCE6F1', 'font_color': '#000000'})
    format_lightblue_black.set_bold()
    format_lightblue_black.set_text_wrap()
    format_lightblue_black.set_border()

    format_header = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#D9D9D9', 'font_color': '#000000'})
    format_header.set_bold()
    format_header.set_text_wrap()

    format_all_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFFFFF', 'font_color': '#FFFFFF'})
    format_all_black = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#000000'})
    format_black_on_white = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#000000', 'font_color': '#FFFFFF'})

    # Format of specific column data
    format_day_data = workbook.add_format({'num_format': 'dd/mm/yyyy', 'valign': 'vcenter'})
    format_day_data.set_align('right')
    format_day_data.set_border()

    format_hour_data = workbook.add_format({'num_format': 'hh:mm:ss', 'valign': 'vcenter'})
    format_hour_data.set_align('right')
    format_hour_data.set_border()

    format_day_hour = workbook.add_format({'num_format': 'dd/mm/yyyy hh:mm:ss', 'valign': 'vcenter'})
    format_day_hour.set_align('right')
    format_day_hour.set_border()

    # Format numbers
    format_number = workbook.add_format({'num_format': '#,##0.00', 'align': 'center', 'valign': 'vcenter'})
    format_number.set_border()

    format_nodecimal = workbook.add_format({'num_format': '0', 'align': 'center', 'valign': 'vcenter'})
    format_nodecimal.set_border()

    format_percentage = workbook.add_format({'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter'})
    format_percentage.set_border()

    format_percentage_good = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#C6EFCE',
         'font_color': '#006100'})
    format_percentage_good.set_border()
    format_percentage_mid = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFEB9C',
         'font_color': '#9C5700'})
    format_percentage_mid.set_border()
    format_percentage_bad = workbook.add_format(
        {'num_format': '0.00%', 'align': 'center', 'valign': 'vcenter', 'bg_color': '#FFC7CE',
         'font_color': '#9C0006'})
    format_percentage_bad.set_border()

    # Format strings
    format_string = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string.set_border()

    format_string_wrapped = workbook.add_format({'align': 'left', 'valign': 'vcenter'})
    format_string_wrapped.set_text_wrap()
    format_string_wrapped.set_border()

    format_string_unlocked = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'locked': False})
    unlocked = workbook.add_format({'locked': False})
    format_string_unlocked.set_border()

    format_string_bold = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold.set_bold()
    format_string_bold.set_border()

    format_string_bold_wrapped = workbook.add_format({'align': 'right', 'valign': 'vcenter'})
    format_string_bold_wrapped.set_bold()
    format_string_bold_wrapped.set_border()
    format_string_bold_wrapped.set_text_wrap()

    format_first_column = workbook.add_format(
        {'align': 'center', 'valign': 'vcenter', 'bg_color': '#F2F2F2', 'font_color': '#000000'})
    format_first_column.set_bold()
    format_first_column.set_border()
    format_first_column.set_text_wrap()
    # </editor-fold>

    # <editor-fold desc="Performance Overview Sheet">
    sheet = "Performance Overview"
    try:
        ws_sheet = workbook.add_worksheet(sheet)
    except (xlsxwriter.exceptions.DuplicateWorksheetName, NameError):
        sheet = sheet + "_new"
        ws_sheet = workbook.add_worksheet(sheet)

    try:
        df_performance = performance_fleet_per_period['choose'].T
    except KeyError:
        df_performance = performance_fleet_per_period['monthly'].T

    sites = list(df_performance.columns)

    start_row_header = 1
    start_row_data = 2
    start_column = 0

    for site in sites:
        level = 0
        start_row_header_str = str(start_row_header)
        start_row_data_str = str(start_row_data)

        performance_site = df_performance.loc[:, [site]].reset_index()
        n_rows_performance = performance_site.shape[0] + 1
        n_columns_performance = performance_site.shape[1]

        df_total = performance_site

        max_rows = n_rows_performance
        n_columns_total = df_total.shape[1]

        width = mf.get_col_widths(df_total)

        print("\n", df_total)

        for i in range(start_column, start_column + n_columns_total):

            header = df_total.columns[i - start_column]

            column_letter = openpyxl.utils.cell.get_column_letter(i + 1)
            header_cell = column_letter + start_row_header_str
            data_cell = column_letter + start_row_data_str
            all_column = column_letter + ':' + column_letter

            # print('Header: ', header, "\n", 'Header cell:', header_cell, "\n", "Data Cell: ", data_cell ,"\n")

            data = list(df_total[header].fillna(""))

            if header == "index":
                to_colapse_column1 = column_letter
                data = [x for x in data if not pd.isnull(x)]
                ws_sheet.write(header_cell, "", format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_lightblue_black)
                if column_letter == "A":
                    ws_sheet.set_column(all_column, 23)
                else:
                    ws_sheet.set_column(all_column, 0)  # ,None,{'level': 1, 'hidden': True})

            elif "LSBP" in header or "Wellington" in header:
                kpis = df_total['index']
                ws_sheet.write(header_cell, header, format_darkblue_white)
                data = [x for x in data if not x == ""]

                for i in range(len(data)):
                    cell = column_letter + str(start_row_data + i)
                    value = data[i]
                    kpi = kpis[i]

                    if "%" in value:
                        value = float(value[:-1]) / 100
                        ws_sheet.write_number(cell, value, format_percentage)
                        if not "PR (%)" in kpi:
                            if "Availability" in kpi:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [{'criteria': '>=', 'type': 'number',
                                                                              'value': 0.97},
                                                                             {'criteria': '<', 'type': 'number',
                                                                              'value': 0.10},
                                                                             {'criteria': '<=', 'type': 'number',
                                                                              'value': 0.10}]})

                            else:
                                ws_sheet.conditional_format(cell, {'type': 'icon_set', 'icon_style': '3_traffic_lights',
                                                                   'icons': [
                                                                       {'criteria': '<=', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>', 'type': 'number', 'value': 0},
                                                                       {'criteria': '>=', 'type': 'number',
                                                                        'value': 0.05}]})


                    else:
                        value = float(value.replace(",", ""))
                        ws_sheet.write_number(cell, value, format_number)

                ws_sheet.set_column(all_column, 16, None)


            else:
                ws_sheet.write(header_cell, header, format_darkblue_white)
                ws_sheet.write_column(data_cell, data, format_string)
                ws_sheet.set_column(all_column, 18, None)  # ,{'level': 1, 'hidden': True})

        level = level + 1
        start_column = start_column + n_columns_total
    # </editor-fold>

    ws_active = workbook.get_worksheet_by_name("Performance Overview")
    ws_active.activate()


    incidents_corrected_period.to_excel(writer_und, sheet_name='Underperformance Report', index=False)

    writer_und.save()

    print('Done')

    return

def get_events_summary_per_fault_component(components_to_analyse, inverter_incidents_site, inverter_operation, df_operation_hours):

    unit_failure_dict = {}
    events_summary_dict = {}
    count = 0

    for unit in inverter_operation.keys():

        # From unit get component, aka, Inv 01.r2 --> Inv 01
        try:
            component = unit.replace(re.search(r'\.r\d*', unit).group(), "")
        except AttributeError:
            component = unit

        # Get unit incidents
        unit_incidents = inverter_incidents_site.loc[inverter_incidents_site['Unit Component'] == unit]
        # print(unit_incidents)

        unit_age = \
        df_operation_hours.loc[df_operation_hours['Timestamp'] == inverter_operation[unit][1]][component].values[0]
        # print(unit, unit_age)

        # Get last time of operation from timestamp, if time empty, look for last datapoint
        changed = False
        while np.isnan(unit_age):
            rounded_incident_time = rounded_incident_time - pd.Timedelta(minutes=15)
            incident_operation_time = \
            df_operation_hours.loc[df_operation_hours['Timestamp'] == rounded_incident_time][component].values[0]
            changed = True
        if changed == True:
            print("Changed rounded time to forward timestamp because backward was NaN, new timestamp: ",
                  rounded_incident_time)

        # From original dataframe, reduce to dataframe with required data
        components_failed = list(set(unit_incidents['Fault Component']))
        events_summary = unit_incidents[['Unit Component', 'Fault Component', 'Event Start Time', 'Operation Time']]
        events_summary['Time to Failure'] = ""
        events_summary['Failure'] = "Yes"

        # Add last entries of dataframe, aka, hours of operation at the last point of analysis
        end_of_analysis_entries = pd.DataFrame({'Unit Component': [unit] * len(components_to_analyse),
                                                'Fault Component': components_to_analyse,
                                                'Event Start Time': [inverter_operation[unit][1]] * len(
                                                    components_to_analyse),
                                                'Operation Time': [unit_age] * len(components_to_analyse),
                                                'Time to Failure': [""] * len(components_to_analyse),
                                                'Failure': ['No'] * len(components_to_analyse)})

        # Get complete events summary
        events_summary = pd.concat([events_summary, end_of_analysis_entries]).sort_values(
            by=['Event Start Time', 'Fault Component']).reset_index(None, drop=True)
        events_summary = events_summary.loc[
            ~(events_summary['Fault Component'] == "Phase Fuse") & ~(events_summary['Fault Component'] == "Unknown")]

        print(events_summary)
        print("\n")

        print(components_failed)
        # Separate multiple components incidents to calculate spare parts
        for failed_component in components_failed:
            if ";" in failed_component:
                incidents_to_split = events_summary.loc[events_summary['Fault Component'] == failed_component]
                index_incidents_to_split = incidents_to_split.index
                actual_components = failed_component.split(';')
                n_repeats = len(actual_components)

                splitted_incidents = pd.concat([incidents_to_split] * len(actual_components))
                splitted_incidents['Fault Component'] = actual_components * len(incidents_to_split)
                splitted_incidents = splitted_incidents.sort_values(
                    by=['Event Start Time', 'Fault Component']).reset_index(None, drop=True)

                events_summary = pd.concat(
                    [events_summary.drop(index=index_incidents_to_split), splitted_incidents]).sort_values(
                    by=['Event Start Time', 'Fault Component']).reset_index(None, drop=True)

                """print(events_summary)
                print(splitted_incidents)
                print(new_events_summary)"""

        # Add time to failure
        for fault_component in components_to_analyse:

            fc_events_summary = events_summary.loc[events_summary['Fault Component'] == fault_component]
            n_incidents = len(fc_events_summary)

            if n_incidents == 1:
                index_of_incident = int(fc_events_summary.index.values)
                events_summary.loc[index_of_incident, "Time to Failure"] = fc_events_summary['Operation Time'][
                    index_of_incident]

            else:
                op_time = list(fc_events_summary['Operation Time'])
                op_time_2 = list(fc_events_summary['Operation Time'])
                op_time_2.insert(0, 0)
                del op_time_2[-1]

                fc_events_summary['Time to Failure'] = [op_time_i - op_time_2_i for op_time_i, op_time_2_i in
                                                        zip(op_time, op_time_2)]

                for index, row in fc_events_summary.iterrows():
                    events_summary.loc[index, "Time to Failure"] = row['Time to Failure']

            fr_calc_events_summary = events_summary.loc[events_summary['Fault Component'] == fault_component]
            n_incidents = len(fr_calc_events_summary.loc[fr_calc_events_summary['Failure'] == 'Yes'])
            n_hours = sum(fr_calc_events_summary['Time to Failure'])
            failure_rate = (n_incidents / n_hours) * 1000

            # print(unit,fault_component, n_incidents, n_hours, failure_rate)

        try:
            all_events_summary = pd.concat([all_events_summary,
                                            events_summary])  # .sort_values(by = ['Event Start Time', 'Fault Component']).reset_index(None, drop=True)
        except NameError:
            all_events_summary = events_summary

        # print(events_summary)

        # print(unit, components_failed)

        unit_failure_dict[unit] = {'Incidents': unit_incidents, 'Unit Age': unit_age, 'Events Summary': events_summary}
        events_summary_dict[unit] = events_summary


    return events_summary_dict, unit_failure_dict, all_events_summary


def get_events_summary_per_failure_mode(components_to_analyse, inverter_incidents_site, inverter_operation, df_operation_hours):

    unit_failure_dict = {}
    events_summary_dict = {}
    count = 0

    for unit in inverter_operation.keys():

        # From unit get component, aka, Inv 01.r2 --> Inv 01
        try:
            component = unit.replace(re.search(r'\.r\d*', unit).group(), "")
        except AttributeError:
            component = unit

        # Get unit incidents
        unit_incidents = inverter_incidents_site.loc[inverter_incidents_site['Unit Component'] == unit]
        # print(unit_incidents)

        unit_age = \
        df_operation_hours.loc[df_operation_hours['Timestamp'] == inverter_operation[unit][1]][component].values[0]
        # print(unit, unit_age)

        # Get last time of operation from timestamp, if time empty, look for last datapoint
        changed = False
        while np.isnan(unit_age):
            rounded_incident_time = rounded_incident_time - pd.Timedelta(minutes=15)
            incident_operation_time = \
            df_operation_hours.loc[df_operation_hours['Timestamp'] == rounded_incident_time][component].values[0]
            changed = True
        if changed == True:
            print("Changed rounded time to forward timestamp because backward was NaN, new timestamp: ",
                  rounded_incident_time)

        # From original dataframe, reduce to dataframe with required data
        components_failed = list(set(unit_incidents['Fault Component']))
        events_summary = unit_incidents[['Unit Component', 'Fault Component', 'Failure Mode', 'Event Start Time', 'Operation Time']]
        events_summary['Time to Failure'] = ""
        events_summary['Failure'] = "Yes"

        # Add last entries of dataframe, aka, hours of operation at the last point of analysis
        end_of_analysis_entries = pd.DataFrame({'Unit Component': [unit] * len(components_to_analyse),
                                                'Fault Component': components_to_analyse,
                                                'Failure Mode': [""] * len(components_to_analyse) ,
                                                'Event Start Time': [inverter_operation[unit][1]] * len(
                                                    components_to_analyse),
                                                'Operation Time': [unit_age] * len(components_to_analyse),
                                                'Time to Failure': [""] * len(components_to_analyse),
                                                'Failure': ['No'] * len(components_to_analyse)})

        # Get complete events summary
        events_summary = pd.concat([events_summary, end_of_analysis_entries]).sort_values(
            by=['Event Start Time', 'Fault Component']).reset_index(None, drop=True)
        events_summary = events_summary.loc[
            ~(events_summary['Fault Component'] == "Phase Fuse") & ~(events_summary['Fault Component'] == "Unknown")]

        print(events_summary)
        print("\n")

        print(components_failed)
        # Separate multiple components incidents to calculate spare parts
        for failed_component in components_failed:
            if ";" in failed_component:
                incidents_to_split = events_summary.loc[events_summary['Fault Component'] == failed_component]
                index_incidents_to_split = incidents_to_split.index
                actual_components = failed_component.split(';')
                n_repeats = len(actual_components)

                splitted_incidents = pd.concat([incidents_to_split] * len(actual_components))
                splitted_incidents['Fault Component'] = actual_components * len(incidents_to_split)
                splitted_incidents = splitted_incidents.sort_values(
                    by=['Event Start Time', 'Fault Component']).reset_index(None, drop=True)

                events_summary = pd.concat(
                    [events_summary.drop(index=index_incidents_to_split), splitted_incidents]).sort_values(
                    by=['Event Start Time', 'Fault Component']).reset_index(None, drop=True)

                """print(events_summary)
                print(splitted_incidents)
                print(new_events_summary)"""

        # Add time to failure
        for fault_component in components_to_analyse:

            fc_events_summary = events_summary.loc[events_summary['Fault Component'] == fault_component]
            n_incidents = len(fc_events_summary)

            if n_incidents == 1:
                index_of_incident = int(fc_events_summary.index.values)
                events_summary.loc[index_of_incident, "Time to Failure"] = fc_events_summary['Operation Time'][
                    index_of_incident]

            else:
                op_time = list(fc_events_summary['Operation Time'])
                op_time_2 = list(fc_events_summary['Operation Time'])
                op_time_2.insert(0, 0)
                del op_time_2[-1]

                fc_events_summary['Time to Failure'] = [op_time_i - op_time_2_i for op_time_i, op_time_2_i in
                                                        zip(op_time, op_time_2)]

                for index, row in fc_events_summary.iterrows():
                    events_summary.loc[index, "Time to Failure"] = row['Time to Failure']

            fr_calc_events_summary = events_summary.loc[events_summary['Fault Component'] == fault_component]
            n_incidents = len(fr_calc_events_summary.loc[fr_calc_events_summary['Failure'] == 'Yes'])
            n_hours = sum(fr_calc_events_summary['Time to Failure'])
            failure_rate = (n_incidents / n_hours) * 1000

            # print(unit,fault_component, n_incidents, n_hours, failure_rate)

        try:
            all_events_summary = pd.concat([all_events_summary,
                                            events_summary])  # .sort_values(by = ['Event Start Time', 'Fault Component']).reset_index(None, drop=True)
        except NameError:
            all_events_summary = events_summary

        # print(events_summary)

        # print(unit, components_failed)

        unit_failure_dict[unit] = {'Incidents': unit_incidents, 'Unit Age': unit_age, 'Events Summary': events_summary}
        events_summary_dict[unit] = events_summary


    return events_summary_dict, unit_failure_dict, all_events_summary
